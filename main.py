from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel, Field
from typing import Literal
import re
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, desc, DECIMAL, SmallInteger, Text, Boolean, Enum, Date, func
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship, joinedload, selectinload
from sqlalchemy.exc import OperationalError
from passlib.context import CryptContext
from jose import JWTError, jwt
import datetime
import io
import pandas as pd
import urllib.parse
import re
import logging
from typing import List, Optional

# 数据库连接配置
import os
from dotenv import load_dotenv

# 加载 .env 文件中的环境变量
load_dotenv()

SQLALCHEMY_DATABASE_URL = os.getenv(
    "DATABASE_URL", 
    "mysql+pymysql://greenhouse_user:change_this_password_in_production@localhost:3306/smart_greenhouse"
)

# 安全配置
SECRET_KEY = os.getenv("SECRET_KEY", os.urandom(32).hex())  # 生产环境必须通过环境变量设置
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# 创建数据库引擎
engine = create_engine(SQLALCHEMY_DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# 密码哈希
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- 数据库模型定义 ---
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True)
    hashed_password = Column(String(255))
    role = Column(String(20))  # 'student', 'teacher', 'admin'
    # 扩展字段
    email = Column(String(100), unique=True, nullable=True)
    real_name = Column(String(50), nullable=True)
    student_id = Column(String(20), unique=True, nullable=True)  # 学号
    teacher_id = Column(String(20), unique=True, nullable=True)  # 工号
    class_id = Column(Integer, ForeignKey("classes.id"), nullable=True)  # 班级 ID
    is_active = Column(Boolean, default=True)  # 账号是否启用
    created_by = Column(Integer, ForeignKey("users.id"), nullable=True)  # 创建人 ID
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class Device(Base):
    __tablename__ = "devices"
    id = Column(Integer, primary_key=True, index=True)
    device_name = Column(String(100), nullable=False)
    status = Column(SmallInteger, default=1)
    last_seen = Column(DateTime, nullable=True)
    pump_state = Column(SmallInteger, default=0)
    fan_state = Column(SmallInteger, default=0)
    light_state = Column(SmallInteger, default=0)
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

class SensorReading(Base):
    __tablename__ = "sensor_readings"
    
    id = Column(Integer, primary_key=True, index=True)
    device_id = Column(Integer, ForeignKey("devices.id"), index=True)
    temp = Column(DECIMAL(5, 2))
    humidity = Column(DECIMAL(5, 2))
    soil_moisture = Column(DECIMAL(5, 2))
    light = Column(DECIMAL(10, 2))
    timestamp = Column(DateTime, default=datetime.datetime.utcnow)

    device = relationship("Device")

# --- 教学内容管理模型 ---
class ContentCategory(Base):
    """教学内容分类表"""
    __tablename__ = "content_categories"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), nullable=False)  # 分类名
    parent_id = Column(Integer, ForeignKey("content_categories.id"), nullable=True)  # 父分类 ID
    description = Column(String(500), nullable=True)
    sort_order = Column(Integer, default=0)  # 排序
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    
    # 自关联：父分类和子分类
    parent = relationship("ContentCategory", remote_side=[id], backref="children")
    # 关联的内容
    contents = relationship("TeachingContent", back_populates="category")

class TeachingContent(Base):
    """教学内容表"""
    __tablename__ = "teaching_contents"
    
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200), nullable=False)  # 标题
    category_id = Column(Integer, ForeignKey("content_categories.id"), nullable=False)  # 分类 ID
    content_type = Column(String(20), default="article")  # article, video, image, pdf
    content = Column(Text, nullable=True)  # 富文本内容
    video_url = Column(String(500), nullable=True)  # 视频链接
    file_path = Column(String(500), nullable=True)  # 文件路径
    cover_image = Column(String(500), nullable=True)  # 封面图
    author_id = Column(Integer, ForeignKey("users.id"), nullable=False)  # 作者
    view_count = Column(Integer, default=0)  # 阅读量
    is_published = Column(Boolean, default=False)  # 是否发布
    published_at = Column(DateTime, nullable=True)  # 发布时间
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    
    # 关联
    category = relationship("ContentCategory", back_populates="contents")
    author = relationship("User")
    # 学习记录
    learning_records = relationship("StudentLearningRecord", back_populates="content")
    # 评论
    comments = relationship("ContentComment", back_populates="content")

class StudentLearningRecord(Base):
    """学生学习记录表"""
    __tablename__ = "student_learning_records"
    
    id = Column(Integer, primary_key=True, index=True)
    student_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    content_id = Column(Integer, ForeignKey("teaching_contents.id"), nullable=False)
    status = Column(String(20), default="not_started")  # not_started, in_progress, completed
    progress_percent = Column(Integer, default=0)  # 学习进度 0-100
    time_spent_seconds = Column(Integer, default=0)  # 学习时长（秒）
    last_accessed = Column(DateTime, default=datetime.datetime.utcnow)
    completed_at = Column(DateTime, nullable=True)
    
    # 关联
    student = relationship("User")
    content = relationship("TeachingContent", back_populates="learning_records")
    
    __table_args__ = (
        # 唯一约束：一个学生对一个内容只有一条记录
        {'sqlite_autoincrement': True}  # 兼容 SQLite（如果需要）
    )

class ContentComment(Base):
    """内容评论/问答表"""
    __tablename__ = "content_comments"
    
    id = Column(Integer, primary_key=True, index=True)
    content_id = Column(Integer, ForeignKey("teaching_contents.id"), nullable=False)
    student_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    comment = Column(Text, nullable=False)  # 评论内容
    teacher_reply = Column(Text, nullable=True)  # 教师回复
    reply_at = Column(DateTime, nullable=True)  # 回复时间
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    
    # 关联
    content = relationship("TeachingContent", back_populates="comments")
    student = relationship("User")

class Class(Base):
    """班级表"""
    __tablename__ = "classes"
    
    id = Column(Integer, primary_key=True, index=True)
    class_name = Column(String(100), nullable=False, index=True)  # 班级名称
    grade = Column(String(20), nullable=True)  # 年级
    teacher_id = Column(Integer, ForeignKey("users.id"), nullable=True)  # 班主任 ID
    description = Column(String(500), nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

class UserOperationLog(Base):
    """用户操作日志表"""
    __tablename__ = "user_operation_logs"

    id = Column(Integer, primary_key=True, index=True)
    operator_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    operation_type = Column(String(20), nullable=False)  # create, update, delete, import, reset_password
    target_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    details = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.datetime.utcnow, index=True)

    # 关联
    operator = relationship("User", foreign_keys=[operator_id])
    target_user = relationship("User", foreign_keys=[target_user_id])


# ==================== 实验报告系统模型 ====================

class Assignment(Base):
    """实验任务表"""
    __tablename__ = "assignments"

    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200), nullable=False)  # 任务标题
    description = Column(Text)  # 任务描述
    device_id = Column(Integer, ForeignKey("devices.id"), index=True)  # 关联设备
    class_id = Column(Integer, ForeignKey("classes.id"), index=True)  # 布置班级
    teacher_id = Column(Integer, ForeignKey("users.id"), index=True)  # 布置教师
    start_date = Column(DateTime)  # 开始时间
    due_date = Column(DateTime)  # 截止时间
    requirement = Column(Text)  # 实验要求
    template = Column(Text)  # 报告模板 (JSON 字符串)
    is_published = Column(Boolean, default=True, index=True)  # 是否发布
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

    # 关联
    device = relationship("Device")
    clas = relationship("Class")
    teacher = relationship("User")
    submissions = relationship("AssignmentSubmission", back_populates="assignment")


class AssignmentSubmission(Base):
    """实验报告提交表"""
    __tablename__ = "assignment_submissions"

    id = Column(Integer, primary_key=True, index=True)
    assignment_id = Column(Integer, ForeignKey("assignments.id"), nullable=False, index=True)
    student_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)

    # 状态
    status = Column(String(20), default="draft", index=True)  # draft, submitted, graded

    # 实验数据
    experiment_date = Column(DateTime)  # 实验日期
    temp_records = Column(Text)  # 温度记录 (JSON 字符串)
    humidity_records = Column(Text)  # 湿度记录 (JSON 字符串)
    soil_moisture_records = Column(Text)  # 土壤湿度记录 (JSON 字符串)
    light_records = Column(Text)  # 光照记录 (JSON 字符串)
    observations = Column(Text)  # 观察记录
    photos = Column(Text)  # 照片路径 (JSON 字符串)
    conclusion = Column(Text)  # 实验结论

    # 评分
    score = Column(DECIMAL(5, 2))  # 分数
    teacher_comment = Column(Text)  # 教师评语
    graded_at = Column(DateTime)  # 批改时间
    graded_by = Column(Integer, ForeignKey("users.id"))  # 批改教师

    submitted_at = Column(DateTime)  # 提交时间
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

    # 关联
    assignment = relationship("Assignment", back_populates="submissions")
    student = relationship("User", foreign_keys=[student_id])
    grader = relationship("User", foreign_keys=[graded_by])


# ==================== 植物生长档案模型 ====================

class PlantProfile(Base):
    """植物档案表"""
    __tablename__ = "plant_profiles"

    id = Column(Integer, primary_key=True, index=True)
    plant_name = Column(String(100), nullable=False)  # 植物名称（如：番茄 01 号）
    species = Column(String(100))  # 品种（如：樱桃番茄）
    class_id = Column(Integer, ForeignKey("classes.id"), index=True)  # 所属班级
    group_id = Column(Integer, ForeignKey("study_groups.id"), index=True)  # 负责小组
    device_id = Column(Integer, ForeignKey("devices.id"), index=True)  # 关联设备
    plant_date = Column(Date)  # 种植日期
    cover_image = Column(String(500))  # 封面图
    status = Column(String(20), default="growing", index=True)  # growing, harvested, withered
    expected_harvest_date = Column(Date)  # 预期收获日期
    description = Column(Text)  # 植物描述
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

    # 关联
    clas = relationship("Class")
    device = relationship("Device")
    growth_records = relationship("GrowthRecord", back_populates="plant")


class GrowthRecord(Base):
    """生长记录表"""
    __tablename__ = "growth_records"

    id = Column(Integer, primary_key=True, index=True)
    plant_id = Column(Integer, ForeignKey("plant_profiles.id", ondelete="CASCADE"), nullable=False, index=True)
    record_date = Column(Date, nullable=False, index=True)  # 记录日期
    stage = Column(String(20))  # seed, sprout, seedling, flowering, fruiting, harvested
    height_cm = Column(DECIMAL(5, 2))  # 高度 (cm)
    leaf_count = Column(Integer)  # 叶片数
    flower_count = Column(Integer)  # 花朵数
    fruit_count = Column(Integer)  # 果实数
    description = Column(Text)  # 观察描述
    photos = Column(Text)  # 照片 (JSON 字符串)
    recorded_by = Column(Integer, ForeignKey("users.id"))  # 记录人
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

    # 关联
    plant = relationship("PlantProfile", back_populates="growth_records")
    recorder = relationship("User")


# ==================== 小组合作学习模型 ====================

class StudyGroup(Base):
    """学习小组表"""
    __tablename__ = "study_groups"

    id = Column(Integer, primary_key=True, index=True)
    group_name = Column(String(100), nullable=False)  # 小组名称
    class_id = Column(Integer, ForeignKey("classes.id"), nullable=False)  # 所属班级
    device_id = Column(Integer, ForeignKey("devices.id"))  # 负责的设备
    description = Column(Text)  # 小组描述
    created_at = Column(DateTime, default=datetime.datetime.utcnow)

    # 关联
    clas = relationship("Class")
    device = relationship("Device")
    members = relationship("GroupMember", back_populates="group")


class GroupMember(Base):
    """小组成员表"""
    __tablename__ = "group_members"

    id = Column(Integer, primary_key=True, index=True)
    group_id = Column(Integer, ForeignKey("study_groups.id"), nullable=False)
    student_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    role = Column(String(20))  # leader, recorder, operator, reporter
    joined_at = Column(DateTime, default=datetime.datetime.utcnow)

    # 关联
    group = relationship("StudyGroup", back_populates="members")
    student = relationship("User")

# --- Pydantic 模型 ---
class Token(BaseModel):
    access_token: str
    token_type: str
    role: str

class UserData(BaseModel):
    username: str
    role: str

class TelemetryData(BaseModel):
    device_id: int
    temp: float
    humidity: float
    soil_moisture: float
    light: float

class TelemetryResponse(TelemetryData):
    id: int
    timestamp: datetime.datetime
    class Config:
        from_attributes = True

class DeviceResponse(BaseModel):
    id: int
    device_name: str
    status: int
    last_seen: Optional[datetime.datetime]
    pump_state: int
    fan_state: int
    light_state: int
    class Config:
        from_attributes = True

class DeviceCreateRequest(BaseModel):
    device_name: str
    status: Optional[int] = 1
    pump_state: Optional[int] = 0
    fan_state: Optional[int] = 0
    light_state: Optional[int] = 0

class ControlRequest(BaseModel):
    pump_state: Optional[int] = None
    fan_state: Optional[int] = None
    light_state: Optional[int] = None

# --- 教学内容管理 Pydantic 模型 ---
class ContentCategoryBase(BaseModel):
    name: str
    description: Optional[str] = None
    parent_id: Optional[int] = None
    sort_order: int = 0

class ContentCategoryCreate(ContentCategoryBase):
    pass

class ContentCategoryResponse(ContentCategoryBase):
    id: int
    created_at: datetime.datetime
    
    class Config:
        from_attributes = True

class ContentCategoryWithChildren(ContentCategoryResponse):
    children: list['ContentCategoryResponse'] = []

class TeachingContentBase(BaseModel):
    title: str
    category_id: int
    content_type: str = "article"
    content: Optional[str] = None
    video_url: Optional[str] = None
    file_path: Optional[str] = None
    cover_image: Optional[str] = None

class TeachingContentCreate(TeachingContentBase):
    is_published: bool = False

class TeachingContentUpdate(BaseModel):
    title: Optional[str] = None
    category_id: Optional[int] = None
    content_type: Optional[str] = None
    content: Optional[str] = None
    video_url: Optional[str] = None
    file_path: Optional[str] = None
    cover_image: Optional[str] = None
    is_published: Optional[bool] = None

class TeachingContentResponse(TeachingContentBase):
    id: int
    author_id: int
    view_count: int
    is_published: bool
    published_at: Optional[datetime.datetime]
    created_at: datetime.datetime
    updated_at: datetime.datetime
    category: Optional[ContentCategoryResponse] = None
    
    class Config:
        from_attributes = True

class TeachingContentDetail(TeachingContentResponse):
    """详细内容响应（包含完整分类信息）"""
    pass

class StudentLearningRecordBase(BaseModel):
    student_id: int
    content_id: int

class StudentLearningRecordCreate(BaseModel):
    status: str = "in_progress"
    progress_percent: int = 0
    time_spent_seconds: int = 0

class StudentLearningRecordResponse(BaseModel):
    id: int
    student_id: int
    content_id: int
    status: str
    progress_percent: int
    time_spent_seconds: int
    last_accessed: datetime.datetime
    completed_at: Optional[datetime.datetime]
    
    class Config:
        from_attributes = True

class ContentCommentBase(BaseModel):
    comment: str

class ContentCommentCreate(ContentCommentBase):
    pass

class ContentCommentResponse(BaseModel):
    id: int
    content_id: int
    student_id: int
    student_name: Optional[str] = None
    comment: str
    teacher_reply: Optional[str] = None
    reply_at: Optional[datetime.datetime] = None
    created_at: datetime.datetime

    class Config:
        from_attributes = True

# --- 用户管理 Pydantic 模型 ---
class UserBase(BaseModel):
    username: str
    role: Literal['student', 'teacher', 'admin']
    email: Optional[str] = None
    real_name: Optional[str] = None

class UserCreate(UserBase):
    password: str
    student_id: Optional[str] = None
    teacher_id: Optional[str] = None
    class_id: Optional[int] = None
    is_active: bool = True

class UserUpdate(BaseModel):
    email: Optional[str] = None
    real_name: Optional[str] = None
    student_id: Optional[str] = None
    teacher_id: Optional[str] = None
    class_id: Optional[int] = None
    is_active: Optional[bool] = None
    role: Optional[str] = None

class UserResponse(UserBase):
    id: int
    student_id: Optional[str] = None
    teacher_id: Optional[str] = None
    class_id: Optional[int] = None
    is_active: bool
    created_at: datetime.datetime
    updated_at: datetime.datetime
    class_name: Optional[str] = None
    
    class Config:
        from_attributes = True

class UserImportResult(BaseModel):
    success_count: int
    error_rows: list
    message: str

class ClassBase(BaseModel):
    class_name: str
    grade: Optional[str] = None
    description: Optional[str] = None
    teacher_id: Optional[int] = None

class ClassCreate(ClassBase):
    pass

class ClassUpdate(BaseModel):
    class_name: Optional[str] = None
    grade: Optional[str] = None
    description: Optional[str] = None
    teacher_id: Optional[int] = None
    is_active: Optional[bool] = None

class ClassResponse(ClassBase):
    id: int
    is_active: bool
    created_at: datetime.datetime
    teacher_name: Optional[str] = None
    student_count: int = 0
    
    class Config:
        from_attributes = True

class UserStats(BaseModel):
    total_users: int
    admin_count: int
    teacher_count: int
    student_count: int
    active_count: int
    inactive_count: int


# ==================== 实验报告系统 Pydantic 模型 ====================

class AssignmentBase(BaseModel):
    title: str
    description: Optional[str] = None
    device_id: Optional[int] = None
    class_id: Optional[int] = None
    start_date: Optional[datetime.datetime] = None
    due_date: Optional[datetime.datetime] = None
    requirement: Optional[str] = None
    template: Optional[str] = None  # JSON 字符串
    is_published: bool = True

class AssignmentCreate(AssignmentBase):
    pass

class AssignmentUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    device_id: Optional[int] = None
    class_id: Optional[int] = None
    start_date: Optional[datetime.datetime] = None
    due_date: Optional[datetime.datetime] = None
    requirement: Optional[str] = None
    template: Optional[str] = None
    is_published: Optional[bool] = None

class AssignmentResponse(AssignmentBase):
    id: int
    teacher_id: int
    teacher_name: Optional[str] = None
    device_name: Optional[str] = None
    class_name: Optional[str] = None
    created_at: datetime.datetime
    submission_count: int = 0

    class Config:
        from_attributes = True

class AssignmentSubmissionBase(BaseModel):
    assignment_id: int
    experiment_date: Optional[datetime.datetime] = None
    observations: Optional[str] = None
    conclusion: Optional[str] = None

class AssignmentSubmissionCreate(AssignmentSubmissionBase):
    temp_records: Optional[str] = None  # JSON 字符串
    humidity_records: Optional[str] = None
    soil_moisture_records: Optional[str] = None
    light_records: Optional[str] = None
    photos: Optional[str] = None  # JSON 字符串

class AssignmentSubmissionUpdate(BaseModel):
    observations: Optional[str] = None
    conclusion: Optional[str] = None
    temp_records: Optional[str] = None
    humidity_records: Optional[str] = None
    soil_moisture_records: Optional[str] = None
    light_records: Optional[str] = None
    photos: Optional[str] = None

class AssignmentSubmissionGrade(BaseModel):
    score: float
    teacher_comment: Optional[str] = None

class AssignmentSubmissionResponse(AssignmentSubmissionBase):
    id: int
    student_id: int
    student_name: Optional[str] = None
    status: str
    temp_records: Optional[str] = None
    humidity_records: Optional[str] = None
    soil_moisture_records: Optional[str] = None
    light_records: Optional[str] = None
    photos: Optional[str] = None
    score: Optional[float] = None
    teacher_comment: Optional[str] = None
    graded_at: Optional[datetime.datetime] = None
    submitted_at: Optional[datetime.datetime] = None
    created_at: datetime.datetime
    updated_at: datetime.datetime

    class Config:
        from_attributes = True


# ==================== 植物生长档案 Pydantic 模型 ====================

class PlantProfileBase(BaseModel):
    plant_name: str
    species: Optional[str] = None
    class_id: Optional[int] = None
    group_id: Optional[int] = None
    device_id: Optional[int] = None
    plant_date: Optional[datetime.date] = None
    cover_image: Optional[str] = None
    status: str = "growing"
    expected_harvest_date: Optional[datetime.date] = None
    description: Optional[str] = None

class PlantProfileCreate(PlantProfileBase):
    pass

class PlantProfileUpdate(BaseModel):
    plant_name: Optional[str] = None
    species: Optional[str] = None
    class_id: Optional[int] = None
    group_id: Optional[int] = None
    device_id: Optional[int] = None
    plant_date: Optional[datetime.date] = None
    cover_image: Optional[str] = None
    status: Optional[str] = None
    expected_harvest_date: Optional[datetime.date] = None
    description: Optional[str] = None

class PlantProfileResponse(PlantProfileBase):
    id: int
    class_name: Optional[str] = None
    device_name: Optional[str] = None
    group_name: Optional[str] = None
    growth_record_count: int = 0
    created_at: datetime.datetime

    class Config:
        from_attributes = True

class GrowthRecordBase(BaseModel):
    plant_id: int
    record_date: datetime.date
    stage: Optional[str] = None  # seed, sprout, seedling, flowering, fruiting, harvested
    height_cm: Optional[float] = None
    leaf_count: Optional[int] = None
    flower_count: Optional[int] = None
    fruit_count: Optional[int] = None
    description: Optional[str] = None
    photos: Optional[str] = None  # JSON 字符串

class GrowthRecordCreate(GrowthRecordBase):
    pass

class GrowthRecordUpdate(BaseModel):
    stage: Optional[str] = None
    height_cm: Optional[float] = None
    leaf_count: Optional[int] = None
    flower_count: Optional[int] = None
    fruit_count: Optional[int] = None
    description: Optional[str] = None
    photos: Optional[str] = None

class GrowthRecordResponse(GrowthRecordBase):
    id: int
    recorder_name: Optional[str] = None
    created_at: datetime.datetime

    class Config:
        from_attributes = True


# ==================== 小组合作学习 Pydantic 模型 ====================

class StudyGroupBase(BaseModel):
    group_name: str
    class_id: int
    device_id: Optional[int] = None
    description: Optional[str] = None

class StudyGroupCreate(StudyGroupBase):
    pass

class StudyGroupUpdate(BaseModel):
    group_name: Optional[str] = None
    class_id: Optional[int] = None
    device_id: Optional[int] = None
    description: Optional[str] = None

class StudyGroupResponse(StudyGroupBase):
    id: int
    class_name: Optional[str] = None
    device_name: Optional[str] = None
    member_count: int = 0
    members: List[dict] = []
    created_at: datetime.datetime

    class Config:
        from_attributes = True

class GroupMemberBase(BaseModel):
    group_id: int
    student_id: int
    role: str  # leader, recorder, operator, reporter

class GroupMemberCreate(GroupMemberBase):
    pass

class GroupMemberResponse(BaseModel):
    id: int
    group_id: int
    student_id: int
    student_name: Optional[str] = None
    role: str
    joined_at: datetime.datetime

    class Config:
        from_attributes = True

# --- FastAPI 实例 ---
app = FastAPI(title="智慧大棚后端 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],  # 限制为前端开发服务器
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- 依赖项 ---
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def create_access_token(data: dict, expires_delta: Optional[datetime.timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.datetime.utcnow() + expires_delta
    else:
        expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise credentials_exception
    return user

async def get_teacher_user(current_user: User = Depends(get_current_user)):
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return current_user

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def get_content_editor(current_user: User = Depends(get_current_user)):
    """验证是否有内容编辑权限（教师或管理员）"""
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="需要教师或管理员权限")
    return current_user

# --- 辅助函数 ---
def log_operation(db: Session, operator_id: int, operation_type: str, target_user_id: Optional[int] = None, details: Optional[str] = None):
    """记录用户操作日志"""
    log = UserOperationLog(
        operator_id=operator_id,
        operation_type=operation_type,
        target_user_id=target_user_id,
        details=details
    )
    db.add(log)
    db.commit()

def validate_password(password: str, strict: bool = False) -> bool:
    """
    验证密码强度
    宽松模式（默认）：至少 6 位
    严格模式（管理员/教师）：至少 8 位，包含大小写字母和数字
    """
    if strict:
        if len(password) < 8:
            return False
        if not re.search(r'[A-Z]', password):
            return False
        if not re.search(r'[a-z]', password):
            return False
        if not re.search(r'\d', password):
            return False
        return True
    else:
        return len(password) >= 6

def validate_username(username: str) -> bool:
    """
    验证用户名格式
    要求：只能包含字母、数字、下划线，3-20 位
    """
    return bool(re.match(r'^[a-zA-Z0-9_]{3,20}$', username))

def validate_email(email: str) -> bool:
    """验证邮箱格式"""
    if not email:
        return True  # 邮箱可选
    return bool(re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email))

# --- 初始化数据 ---
@app.on_event("startup")
def startup_event():
    # 确保表存在
    Base.metadata.create_all(bind=engine)
    
    # 初始化用户
    db = SessionLocal()
    users = [
        ("admin", "admin", "admin"),  # 密码: admin
        ("teacher", "admin", "teacher"),  # 密码: admin (演示用，生产改)
        ("student", "admin", "student")   # 密码: admin
    ]
    for username, password, role in users:
        if not db.query(User).filter(User.username == username).first():
            hashed_pw = pwd_context.hash(password)
            db.add(User(username=username, hashed_password=hashed_pw, role=role))
    
    # 初始化默认设备
    if not db.query(Device).first():
        default_device = Device(
            device_name='Default greenhouse',
            status=1,
            pump_state=0,
            fan_state=0,
            light_state=0
        )
        db.add(default_device)
    
    # 初始化默认教学内容分类
    if not db.query(ContentCategory).first():
        # 一级分类
        categories = [
            ContentCategory(name="农作物习性", description="各种农作物的生长习性和环境需求", sort_order=1),
            ContentCategory(name="植物百科", description="植物分类和特征介绍", sort_order=2),
            ContentCategory(name="自然科学", description="自然科学基础知识", sort_order=3),
            ContentCategory(name="实验指导", description="实验步骤和操作指南", sort_order=4),
        ]
        for cat in categories:
            db.add(cat)
        db.commit()

        # 获取刚添加的分类 ID
        db.refresh(categories[0])
        db.refresh(categories[1])
        db.refresh(categories[2])
        db.refresh(categories[3])

        # 二级分类（农作物习性的子分类）
        sub_categories = [
            ContentCategory(name="温度需求", parent_id=categories[0].id, description="不同作物的温度适应性", sort_order=1),
            ContentCategory(name="光照需求", parent_id=categories[0].id, description="光照对作物生长的影响", sort_order=2),
            ContentCategory(name="水分需求", parent_id=categories[0].id, description="作物灌溉和水分管理", sort_order=3),
            ContentCategory(name="土壤要求", parent_id=categories[0].id, description="土壤类型和肥料需求", sort_order=4),
        ]
        for sub_cat in sub_categories:
            db.add(sub_cat)

        db.commit()

    # 初始化默认班级
    if not db.query(Class).first():
        # 获取教师用户（用于分配为班主任）
        teacher_user = db.query(User).filter(User.role == 'teacher').first()
        teacher_id = teacher_user.id if teacher_user else None
        
        classes = [
            Class(class_name="三年级 1 班", grade="三年级", description="2024 级 1 班", teacher_id=teacher_id),
            Class(class_name="三年级 2 班", grade="三年级", description="2024 级 2 班", teacher_id=teacher_id),
            Class(class_name="四年级 1 班", grade="四年级", description="2023 级 1 班", teacher_id=teacher_id),
        ]
        for cls in classes:
            db.add(cls)
        db.commit()

    db.commit()
    db.close()

@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not pwd_context.verify(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer", "role": user.role}

@app.post("/api/telemetry")
async def receive_telemetry(data: TelemetryData, db: Session = Depends(get_db)):
    try:
        device = db.query(Device).filter(Device.id == data.device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="Device not found")

        new_reading = SensorReading(
            device_id=data.device_id,
            temp=data.temp,
            humidity=data.humidity,
            soil_moisture=data.soil_moisture,
            light=data.light
        )
        db.add(new_reading)
        
        device.last_seen = datetime.datetime.utcnow()
        device.status = 1 
        
        db.commit()
        db.refresh(new_reading)
        
        return {
            "status": "success", 
            "id": new_reading.id,
            "commands": {
                "pump": device.pump_state,
                "fan": device.fan_state,
                "light": device.light_state
            }
        }
    except OperationalError:
        raise HTTPException(status_code=500, detail="Database connection failed")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/history/{device_id}", response_model=List[TelemetryResponse])
async def get_history(device_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    readings = db.query(SensorReading)\
        .filter(SensorReading.device_id == device_id)\
        .order_by(desc(SensorReading.timestamp))\
        .limit(20)\
        .all()
    return readings

@app.get("/api/devices", response_model=List[DeviceResponse])
async def get_devices(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    devices = db.query(Device).all()
    return devices

@app.post("/api/devices", response_model=DeviceResponse)
async def create_device(device_data: DeviceCreateRequest, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    new_device = Device(
        device_name=device_data.device_name,
        status=device_data.status,
        pump_state=device_data.pump_state,
        fan_state=device_data.fan_state,
        light_state=device_data.light_state
    )
    db.add(new_device)
    db.commit()
    db.refresh(new_device)
    return new_device

@app.post("/api/control/{device_id}")
async def control_device(device_id: int, data: ControlRequest, db: Session = Depends(get_db), current_user: User = Depends(get_teacher_user)):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    if data.pump_state is not None:
        device.pump_state = data.pump_state
    if data.fan_state is not None:
        device.fan_state = data.fan_state
    if data.light_state is not None:
        device.light_state = data.light_state
    
    db.commit()
    db.refresh(device)
    return {"status": "success", "device_id": device_id}

class ExportRequest(BaseModel):
    device_id: int
    start_date: str  # YYYY-MM-DD format
    end_date: str    # YYYY-MM-DD format

@app.post("/api/telemetry/export")
async def export_telemetry(
    request: ExportRequest,
    export_format: str = "csv",  # csv or xlsx
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出传感器数据历史记录"""
    try:
        # 验证设备是否存在
        device = db.query(Device).filter(Device.id == request.device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="Device not found")

        # 权限验证
        # 当前学生可以导出所有设备数据（教学用途）
        # TODO: 实现班级系统后，限制学生只能导出自己班级/分组的设备
        # if current_user.role == 'student':
        #     class_membership = db.query(ClassMember).filter(
        #         ClassMember.student_id == current_user.id,
        #         ClassMember.device_id == request.device_id
        #     ).first()
        #     if not class_membership:
        #         raise HTTPException(status_code=403, detail="No permission to export this device's data")

        # 解析并验证日期
        try:
            start_date = datetime.datetime.strptime(request.start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(request.end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，请使用 YYYY-MM-DD 格式")

        # 验证日期范围
        now = datetime.datetime.now()
        if start_date > end_date:
            raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
        if end_date > now:
            raise HTTPException(status_code=400, detail="结束日期不能是未来时间")
        if (end_date - start_date).days > 31:
            raise HTTPException(status_code=400, detail="日期范围不能超过 31 天")

        # 查询数据（包含结束日期当天）
        end_date_inclusive = end_date + datetime.timedelta(days=1)
        readings = db.query(SensorReading)\
            .filter(SensorReading.device_id == request.device_id)\
            .filter(SensorReading.timestamp >= start_date)\
            .filter(SensorReading.timestamp < end_date_inclusive)\
            .order_by(SensorReading.timestamp)\
            .all()

        if not readings:
            raise HTTPException(status_code=404, detail="所选日期范围内没有数据")

        # 转换为 DataFrame
        data = []
        for r in readings:
            data.append({
                "时间": r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "设备 ID": r.device_id,
                "设备名称": device.device_name,
                "温度 (°C)": float(r.temp) if r.temp is not None else "",
                "湿度 (%)": float(r.humidity) if r.humidity is not None else "",
                "土壤湿度 (%)": float(r.soil_moisture) if r.soil_moisture is not None else "",
                "光照强度 (Lx)": float(r.light) if r.light is not None else ""
            })

        df = pd.DataFrame(data)

        # 清理文件名中的特殊字符（防止路径遍历）
        safe_device_name = re.sub(r'[^\w\s-]', '', device.device_name).strip()
        safe_device_name = re.sub(r'\s+', '_', safe_device_name)  # 空格转下划线

        # 导出文件
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        # 记录导出日志（审计用途）
        logging.info(f"User {current_user.username} exported {export_format.upper()} data for device {device.id} ({device.device_name})")

        if export_format == "xlsx":
            # Excel 导出
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='SensorData')
            output.seek(0)

            filename = f"sensor_data_{safe_device_name}_{timestamp}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            # CSV 导出
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)  # 重置指针到开始

            filename = f"sensor_data_{safe_device_name}_{timestamp}.csv"
            media_type = "text/csv"

        # URL encode filename for Content-Disposition header
        encoded_filename = urllib.parse.quote(filename)

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Export failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出失败：{str(e)}")

@app.get("/")
async def root():
    return {"message": "智慧大棚后端服务运行中..."}


# ==================== 公开 API（无需认证） ====================

@app.get("/api/public/display")
async def get_display_data(db: Session = Depends(get_db)):
    """
    公开数据大屏接口 - 无需认证
    返回脱敏的实时数据，适合公开展示
    """
    # 获取默认设备
    device = db.query(Device).filter(Device.id == 1).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    # 获取最新传感器读数
    latest_reading = db.query(SensorReading)\
        .filter(SensorReading.device_id == device.id)\
        .order_by(desc(SensorReading.timestamp))\
        .first()
    
    # 获取生长中的植物
    plants = db.query(PlantProfile).filter(PlantProfile.status == "growing").limit(4).all()
    
    # 获取最近的生长记录
    recent_records = db.query(GrowthRecord)\
        .join(PlantProfile)\
        .order_by(desc(GrowthRecord.record_date))\
        .limit(5)\
        .all()
    
    # 格式化响应
    records_data = []
    for record in recent_records:
        records_data.append({
            "id": record.id,
            "plant_name": record.plant.plant_name if record.plant else "未知",
            "record_date": record.record_date.isoformat() if record.record_date else None,
            "stage": record.stage,
            "height_cm": float(record.height_cm) if record.height_cm else None,
            "leaf_count": record.leaf_count
        })
    
    return {
        "device": {
            "id": device.id,
            "name": device.device_name,
            "status": device.status,
            "pump_state": device.pump_state,
            "fan_state": device.fan_state,
            "light_state": device.light_state
        },
        "telemetry": {
            "temp": float(latest_reading.temp) if latest_reading and latest_reading.temp else None,
            "humidity": float(latest_reading.humidity) if latest_reading and latest_reading.humidity else None,
            "soil_moisture": float(latest_reading.soil_moisture) if latest_reading and latest_reading.soil_moisture else None,
            "light": float(latest_reading.light) if latest_reading and latest_reading.light else None,
            "timestamp": latest_reading.timestamp.isoformat() if latest_reading else None
        },
        "plants": [
            {
                "id": p.id,
                "name": p.plant_name,
                "species": p.species,
                "status": p.status
            }
            for p in plants
        ],
        "recent_records": records_data
    }

# ==================== 教学内容管理 API ====================

# ---------------- 分类管理 ----------------
@app.get("/api/content/categories", response_model=List[ContentCategoryResponse])
async def get_categories(db: Session = Depends(get_db)):
    """获取所有分类（按排序排序）"""
    categories = db.query(ContentCategory).filter(
        ContentCategory.parent_id == None
    ).order_by(ContentCategory.sort_order).all()
    return categories

@app.get("/api/content/categories/tree", response_model=List[ContentCategoryWithChildren])
async def get_categories_tree(db: Session = Depends(get_db)):
    """获取分类树形结构"""
    parent_categories = db.query(ContentCategory).filter(
        ContentCategory.parent_id == None
    ).order_by(ContentCategory.sort_order).all()
    
    result = []
    for parent in parent_categories:
        parent_dict = ContentCategoryWithChildren(
            id=parent.id,
            name=parent.name,
            description=parent.description,
            parent_id=parent.parent_id,
            sort_order=parent.sort_order,
            created_at=parent.created_at,
            children=[]
        )
        # 获取子分类
        children = db.query(ContentCategory).filter(
            ContentCategory.parent_id == parent.id
        ).order_by(ContentCategory.sort_order).all()
        parent_dict.children = [
            ContentCategoryResponse(
                id=child.id,
                name=child.name,
                description=child.description,
                parent_id=child.parent_id,
                sort_order=child.sort_order,
                created_at=child.created_at
            )
            for child in children
        ]
        result.append(parent_dict)
    
    return result

@app.post("/api/content/categories", response_model=ContentCategoryResponse)
async def create_category(
    category: ContentCategoryCreate,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """创建分类（教师/管理员）"""
    db_category = ContentCategory(**category.dict())
    db.add(db_category)
    db.commit()
    db.refresh(db_category)
    return db_category

@app.put("/api/content/categories/{category_id}", response_model=ContentCategoryResponse)
async def update_category(
    category_id: int,
    category: ContentCategoryCreate,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """更新分类（教师/管理员）"""
    db_category = db.query(ContentCategory).filter(ContentCategory.id == category_id).first()
    if not db_category:
        raise HTTPException(status_code=404, detail="分类不存在")
    
    for key, value in category.dict().items():
        setattr(db_category, key, value)
    
    db.commit()
    db.refresh(db_category)
    return db_category

@app.delete("/api/content/categories/{category_id}")
async def delete_category(
    category_id: int,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """删除分类（仅管理员）"""
    db_category = db.query(ContentCategory).filter(ContentCategory.id == category_id).first()
    if not db_category:
        raise HTTPException(status_code=404, detail="分类不存在")
    
    # 检查是否有子分类或内容
    has_children = db.query(ContentCategory).filter(ContentCategory.parent_id == category_id).first()
    if has_children:
        raise HTTPException(status_code=400, detail="请先删除子分类")
    
    has_contents = db.query(TeachingContent).filter(TeachingContent.category_id == category_id).first()
    if has_contents:
        raise HTTPException(status_code=400, detail="请先删除该分类下的内容")
    
    db.delete(db_category)
    db.commit()
    return {"message": "分类已删除"}

# ---------------- 内容管理 ----------------
@app.get("/api/content/contents", response_model=List[TeachingContentResponse])
async def get_contents(
    category_id: Optional[int] = None,
    content_type: Optional[str] = None,
    is_published: Optional[bool] = True,
    search: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取内容列表（支持筛选和搜索）"""
    query = db.query(TeachingContent)

    # 非教师/管理员只能查看已发布的内容
    if current_user.role not in ['teacher', 'admin']:
        is_published = True

    if category_id:
        query = query.filter(TeachingContent.category_id == category_id)
    if content_type:
        query = query.filter(TeachingContent.content_type == content_type)
    if is_published is not None:
        query = query.filter(TeachingContent.is_published == is_published)
    
    # 搜索功能：支持标题和内容搜索
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (TeachingContent.title.like(search_pattern)) |
            (TeachingContent.content.like(search_pattern))
        )

    contents = query.order_by(desc(TeachingContent.created_at)).offset(offset).limit(limit).all()
    return contents

@app.get("/api/content/contents/{content_id}", response_model=TeachingContentDetail)
async def get_content(
    content_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取内容详情"""
    content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    # 非教师/管理员只能查看已发布的内容
    if current_user.role not in ['teacher', 'admin'] and not content.is_published:
        raise HTTPException(status_code=403, detail="无权查看未发布的内容")
    
    # 增加阅读量
    content.view_count += 1
    db.commit()
    db.refresh(content)
    
    return content

@app.post("/api/content/contents", response_model=TeachingContentResponse)
async def create_content(
    content: TeachingContentCreate,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """创建内容（教师/管理员）"""
    db_content = TeachingContent(
        **content.dict(),
        author_id=current_user.id
    )
    if content.is_published:
        db_content.published_at = datetime.datetime.utcnow()
    
    db.add(db_content)
    db.commit()
    db.refresh(db_content)
    return db_content

@app.put("/api/content/contents/{content_id}", response_model=TeachingContentResponse)
async def update_content(
    content_id: int,
    content_update: TeachingContentUpdate,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """更新内容（教师/管理员）"""
    db_content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not db_content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    # 教师只能编辑自己创建的内容
    if current_user.role != 'admin' and db_content.author_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能编辑自己创建的内容")
    
    update_data = content_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_content, key, value)
    
    # 如果是发布操作，设置发布时间
    if update_data.get('is_published') and not db_content.is_published:
        db_content.published_at = datetime.datetime.utcnow()
    
    db.commit()
    db.refresh(db_content)
    return db_content

@app.delete("/api/content/contents/{content_id}")
async def delete_content(
    content_id: int,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """删除内容（教师/管理员）"""
    db_content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not db_content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    # 教师只能删除自己创建的内容
    if current_user.role != 'admin' and db_content.author_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能删除自己创建的内容")
    
    db.delete(db_content)
    db.commit()
    return {"message": "内容已删除"}

@app.post("/api/content/contents/{content_id}/publish")
async def publish_content(
    content_id: int,
    current_user: User = Depends(get_content_editor),
    db: Session = Depends(get_db)
):
    """发布/取消发布内容"""
    db_content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not db_content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    db_content.is_published = not db_content.is_published
    if db_content.is_published:
        db_content.published_at = datetime.datetime.utcnow()
    else:
        db_content.published_at = None
    
    db.commit()
    return {"message": f"内容已{'发布' if db_content.is_published else '取消发布'}"}

# ---------------- 学习记录 ----------------
@app.get("/api/content/my-learning", response_model=List[StudentLearningRecordResponse])
async def get_my_learning(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取我的学习记录（学生）"""
    records = db.query(StudentLearningRecord).filter(
        StudentLearningRecord.student_id == current_user.id
    ).all()
    return records

@app.post("/api/content/contents/{content_id}/start", response_model=StudentLearningRecordResponse)
async def start_learning(
    content_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """开始学习（创建或更新学习记录）"""
    # 检查内容是否存在
    content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    # 查找或创建学习记录
    record = db.query(StudentLearningRecord).filter(
        StudentLearningRecord.student_id == current_user.id,
        StudentLearningRecord.content_id == content_id
    ).first()
    
    if record:
        record.status = "in_progress"
        record.last_accessed = datetime.datetime.utcnow()
    else:
        record = StudentLearningRecord(
            student_id=current_user.id,
            content_id=content_id,
            status="in_progress"
        )
        db.add(record)
    
    db.commit()
    db.refresh(record)
    return record

@app.post("/api/content/contents/{content_id}/complete", response_model=StudentLearningRecordResponse)
async def complete_learning(
    content_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """完成学习"""
    record = db.query(StudentLearningRecord).filter(
        StudentLearningRecord.student_id == current_user.id,
        StudentLearningRecord.content_id == content_id
    ).first()
    
    if not record:
        record = StudentLearningRecord(
            student_id=current_user.id,
            content_id=content_id,
            status="completed",
            completed_at=datetime.datetime.utcnow()
        )
        db.add(record)
    else:
        record.status = "completed"
        record.completed_at = datetime.datetime.utcnow()
        record.progress_percent = 100
    
    db.commit()
    db.refresh(record)
    return record

@app.put("/api/content/contents/{content_id}/progress")
async def update_progress(
    content_id: int,
    progress: int,
    time_spent: int = 0,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """更新学习进度"""
    record = db.query(StudentLearningRecord).filter(
        StudentLearningRecord.student_id == current_user.id,
        StudentLearningRecord.content_id == content_id
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="学习记录不存在")
    
    record.progress_percent = min(100, max(0, progress))
    record.time_spent_seconds = time_spent
    record.last_accessed = datetime.datetime.utcnow()
    
    if record.progress_percent >= 100:
        record.status = "completed"
        record.completed_at = datetime.datetime.utcnow()
    
    db.commit()
    return {"message": "进度已更新", "progress": record.progress_percent}

# ---------------- 评论/问答 ----------------
@app.get("/api/content/contents/{content_id}/comments", response_model=List[ContentCommentResponse])
async def get_comments(
    content_id: int,
    db: Session = Depends(get_db)
):
    """获取内容的评论列表"""
    comments = db.query(ContentComment).filter(
        ContentComment.content_id == content_id
    ).order_by(ContentComment.created_at).all()
    
    # 添加学生姓名
    result = []
    for comment in comments:
        student = db.query(User).filter(User.id == comment.student_id).first()
        comment_dict = ContentCommentResponse(
            id=comment.id,
            content_id=comment.content_id,
            student_id=comment.student_id,
            student_name=student.username if student else None,
            comment=comment.comment,
            teacher_reply=comment.teacher_reply,
            reply_at=comment.reply_at,
            created_at=comment.created_at
        )
        result.append(comment_dict)
    
    return result

@app.post("/api/content/contents/{content_id}/comments", response_model=ContentCommentResponse)
async def add_comment(
    content_id: int,
    comment: ContentCommentCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """添加评论（学生）"""
    # 检查内容是否存在
    content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    db_comment = ContentComment(
        content_id=content_id,
        student_id=current_user.id,
        comment=comment.comment
    )
    db.add(db_comment)
    db.commit()
    db.refresh(db_comment)
    
    return ContentCommentResponse(
        id=db_comment.id,
        content_id=db_comment.content_id,
        student_id=db_comment.student_id,
        student_name=current_user.username,
        comment=db_comment.comment,
        teacher_reply=db_comment.teacher_reply,
        reply_at=db_comment.reply_at,
        created_at=db_comment.created_at
    )

@app.put("/api/content/comments/{comment_id}", response_model=ContentCommentResponse)
async def reply_comment(
    comment_id: int,
    reply: str,
    current_user: User = Depends(get_teacher_user),
    db: Session = Depends(get_db)
):
    """回复评论（教师/管理员）"""
    db_comment = db.query(ContentComment).filter(ContentComment.id == comment_id).first()
    if not db_comment:
        raise HTTPException(status_code=404, detail="评论不存在")
    
    db_comment.teacher_reply = reply
    db_comment.reply_at = datetime.datetime.utcnow()

    db.commit()
    db.refresh(db_comment)

    return ContentCommentResponse(
        id=db_comment.id,
        content_id=db_comment.content_id,
        student_id=db_comment.student_id,
        student_name=db_comment.student.username if db_comment.student else None,
        comment=db_comment.comment,
        teacher_reply=db_comment.teacher_reply,
        reply_at=db_comment.reply_at,
        created_at=db_comment.created_at
    )

# ---------------- 学习统计（教师端） ----------------
class LearningStats(BaseModel):
    """学习统计数据"""
    total_students: int
    total_contents: int
    total_learning_records: int
    completed_count: int
    in_progress_count: int
    not_started_count: int
    completion_rate: float  # 完成率
    average_progress: float  # 平均进度

class StudentProgress(BaseModel):
    """学生学习进度"""
    student_id: int
    student_name: str
    total_contents: int
    completed_count: int
    in_progress_count: int
    completion_rate: float
    total_time_spent: int  # 总学习时长（秒）

@app.get("/api/content/stats/overview", response_model=LearningStats)
async def get_learning_stats(
    current_user: User = Depends(get_teacher_user),
    db: Session = Depends(get_db)
):
    """获取学习统计概览（教师/管理员）"""
    # 学生总数
    total_students = db.query(User).filter(User.role == 'student').count()
    
    # 内容总数
    total_contents = db.query(TeachingContent).filter(TeachingContent.is_published == True).count()
    
    # 学习记录统计
    total_records = db.query(StudentLearningRecord).count()
    completed = db.query(StudentLearningRecord).filter(StudentLearningRecord.status == 'completed').count()
    in_progress = db.query(StudentLearningRecord).filter(StudentLearningRecord.status == 'in_progress').count()
    not_started = total_records - completed - in_progress
    
    # 计算完成率和平均进度
    completion_rate = (completed / total_contents * 100) if total_contents > 0 else 0
    avg_progress_result = db.query(
        (db.func.sum(StudentLearningRecord.progress_percent) / db.func.count(StudentLearningRecord)).label('avg')
    ).filter(StudentLearningRecord.progress_percent > 0).first()
    average_progress = avg_progress_result.avg if avg_progress_result and avg_progress_result.avg else 0
    
    return LearningStats(
        total_students=total_students,
        total_contents=total_contents,
        total_learning_records=total_records,
        completed_count=completed,
        in_progress_count=in_progress,
        not_started_count=not_started,
        completion_rate=round(completion_rate, 2),
        average_progress=round(average_progress, 2)
    )

@app.get("/api/content/stats/students", response_model=List[StudentProgress])
async def get_students_progress(
    current_user: User = Depends(get_teacher_user),
    db: Session = Depends(get_db)
):
    """获取所有学生学习进度（教师/管理员）"""
    students = db.query(User).filter(User.role == 'student').all()
    published_contents = db.query(TeachingContent).filter(TeachingContent.is_published == True).count()
    
    result = []
    for student in students:
        records = db.query(StudentLearningRecord).filter(
            StudentLearningRecord.student_id == student.id
        ).all()
        
        completed = sum(1 for r in records if r.status == 'completed')
        in_progress = sum(1 for r in records if r.status == 'in_progress')
        total_time = sum(r.time_spent_seconds for r in records)
        
        completion_rate = (completed / published_contents * 100) if published_contents > 0 else 0
        
        result.append(StudentProgress(
            student_id=student.id,
            student_name=student.username,
            total_contents=published_contents,
            completed_count=completed,
            in_progress_count=in_progress,
            completion_rate=round(completion_rate, 2),
            total_time_spent=total_time
        ))
    
    return result

@app.get("/api/content/stats/content/{content_id}", response_model=List[StudentProgress])
async def get_content_learning_stats(
    content_id: int,
    current_user: User = Depends(get_teacher_user),
    db: Session = Depends(get_db)
):
    """获取特定内容的学习统计（教师/管理员）"""
    content = db.query(TeachingContent).filter(TeachingContent.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="内容不存在")
    
    records = db.query(StudentLearningRecord).filter(
        StudentLearningRecord.content_id == content_id
    ).all()
    
    result = []
    for record in records:
        student = db.query(User).filter(User.id == record.student_id).first()
        if student:
            result.append(StudentProgress(
                student_id=student.id,
                student_name=student.username,
                total_contents=1,
                completed_count=1 if record.status == 'completed' else 0,
                in_progress_count=1 if record.status == 'in_progress' else 0,
                completion_rate=100 if record.status == 'completed' else 0,
                total_time_spent=record.time_spent_seconds
            ))
    
    return result

# ==================== 用户管理 API ====================

# ---------------- 用户管理 ----------------
@app.get("/api/users", response_model=List[UserResponse])
async def get_users(
    role: Optional[str] = None,
    class_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """获取用户列表（支持搜索/筛选）"""
    query = db.query(User)
    
    if role:
        query = query.filter(User.role == role)
    if class_id:
        query = query.filter(User.class_id == class_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (User.username.like(search_pattern)) |
            (User.real_name.like(search_pattern)) |
            (User.student_id.like(search_pattern))
        )
    
    users = query.order_by(desc(User.created_at)).offset(offset).limit(limit).all()
    
    # 添加班级名称
    result = []
    for user in users:
        class_name = None
        if user.class_id:
            cls = db.query(Class).filter(Class.id == user.class_id).first()
            if cls:
                class_name = cls.class_name
        user_dict = UserResponse(
            id=user.id,
            username=user.username,
            role=user.role,
            email=user.email,
            real_name=user.real_name,
            student_id=user.student_id,
            teacher_id=user.teacher_id,
            class_id=user.class_id,
            is_active=user.is_active,
            created_at=user.created_at,
            updated_at=user.updated_at,
            class_name=class_name
        )
        result.append(user_dict)
    
    return result

@app.post("/api/users", response_model=UserResponse)
async def create_user(
    user: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """创建用户（管理员）"""
    # 验证用户名格式
    if not validate_username(user.username):
        raise HTTPException(
            status_code=400,
            detail="用户名格式无效：只能包含字母、数字、下划线，长度 3-20 位"
        )

    # 验证密码强度（管理员/教师需要严格密码，学生可用简单密码）
    strict_mode = user.role in ['admin', 'teacher']
    password_rule = "至少 8 位，包含大小写字母和数字" if strict_mode else "至少 6 位"
    if not validate_password(user.password, strict=strict_mode):
        raise HTTPException(
            status_code=400,
            detail=f"密码强度不足：{password_rule}"
        )

    # 验证邮箱格式
    if user.email and not validate_email(user.email):
        raise HTTPException(status_code=400, detail="邮箱格式无效")

    # 检查用户名是否已存在
    existing = db.query(User).filter(User.username == user.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")

    # 检查学号/工号是否重复
    if user.student_id:
        if db.query(User).filter(User.student_id == user.student_id).first():
            raise HTTPException(status_code=400, detail="学号已存在")
    if user.teacher_id:
        if db.query(User).filter(User.teacher_id == user.teacher_id).first():
            raise HTTPException(status_code=400, detail="工号已存在")

    # 创建用户
    hashed_password = pwd_context.hash(user.password)
    db_user = User(
        username=user.username,
        hashed_password=hashed_password,
        role=user.role,
        real_name=user.real_name,
        email=user.email,
        student_id=user.student_id,
        teacher_id=user.teacher_id,
        class_id=user.class_id,
        is_active=user.is_active,
        created_by=current_user.id
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)

    # 记录日志
    log_operation(db, current_user.id, 'create', db_user.id, f"创建用户：{user.username} ({user.role})")

    class_name = None
    if db_user.class_id:
        cls = db.query(Class).filter(Class.id == db_user.class_id).first()
        if cls:
            class_name = cls.class_name

    return UserResponse(
        id=db_user.id,
        username=db_user.username,
        role=db_user.role,
        email=db_user.email,
        real_name=db_user.real_name,
        student_id=db_user.student_id,
        teacher_id=db_user.teacher_id,
        class_id=db_user.class_id,
        is_active=db_user.is_active,
        created_at=db_user.created_at,
        updated_at=db_user.updated_at,
        class_name=class_name
    )

@app.get("/api/users/{user_id}", response_model=UserResponse)
async def get_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    """获取用户详情"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    class_name = None
    if user.class_id:
        cls = db.query(Class).filter(Class.id == user.class_id).first()
        if cls:
            class_name = cls.class_name
    
    return UserResponse(
        id=user.id,
        username=user.username,
        role=user.role,
        email=user.email,
        real_name=user.real_name,
        student_id=user.student_id,
        teacher_id=user.teacher_id,
        class_id=user.class_id,
        is_active=user.is_active,
        created_at=user.created_at,
        updated_at=user.updated_at,
        class_name=class_name
    )

@app.put("/api/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """更新用户信息"""
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    update_data = user_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_user, key, value)
    
    db.commit()
    db.refresh(db_user)
    
    # 记录日志
    log_operation(db, current_user.id, 'update', db_user.id, f"更新用户：{db_user.username}")
    
    class_name = None
    if db_user.class_id:
        cls = db.query(Class).filter(Class.id == db_user.class_id).first()
        if cls:
            class_name = cls.class_name
    
    return UserResponse(
        id=db_user.id,
        username=db_user.username,
        role=db_user.role,
        email=db_user.email,
        real_name=db_user.real_name,
        student_id=db_user.student_id,
        teacher_id=db_user.teacher_id,
        class_id=db_user.class_id,
        is_active=db_user.is_active,
        created_at=db_user.created_at,
        updated_at=db_user.updated_at,
        class_name=class_name
    )

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    """删除用户"""
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    username = db_user.username
    db.delete(db_user)
    db.commit()
    
    # 记录日志
    log_operation(db, current_user.id, 'delete', user_id, f"删除用户：{username}")
    
    return {"message": "用户已删除"}

@app.post("/api/users/{user_id}/reset-password")
async def reset_password(
    user_id: int,
    new_password: str = Query(..., min_length=8),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """重置密码"""
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 验证密码强度
    if not validate_password(new_password):
        raise HTTPException(
            status_code=400, 
            detail="密码强度不足：必须至少 8 位，且包含大小写字母和数字"
        )

    db_user.hashed_password = pwd_context.hash(new_password)
    db.commit()

    # 记录日志
    log_operation(db, current_user.id, 'reset_password', user_id, f"重置用户密码：{db_user.username}")

    return {"message": "密码已重置"}

@app.post("/api/users/{user_id}/toggle-active")
async def toggle_user_active(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """启用/禁用账号"""
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    db_user.is_active = not db_user.is_active
    db.commit()
    
    # 记录日志
    status_text = "启用" if db_user.is_active else "禁用"
    log_operation(db, current_user.id, 'update', user_id, f"{status_text}用户：{db_user.username}")
    
    return {"message": f"账号已{status_text}", "is_active": db_user.is_active}

@app.get("/api/stats/users", response_model=UserStats)
async def get_user_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """获取用户统计"""
    total = db.query(User).count()
    admin_count = db.query(User).filter(User.role == 'admin').count()
    teacher_count = db.query(User).filter(User.role == 'teacher').count()
    student_count = db.query(User).filter(User.role == 'student').count()
    active = db.query(User).filter(User.is_active == True).count()
    inactive = db.query(User).filter(User.is_active == False).count()
    
    return UserStats(
        total_users=total,
        admin_count=admin_count,
        teacher_count=teacher_count,
        student_count=student_count,
        active_count=active,
        inactive_count=inactive
    )

@app.post("/api/users/batch-create")
async def batch_create_users(
    users: List[UserCreate],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """批量创建用户"""
    success_count = 0
    error_users = []
    
    for user_data in users:
        try:
            # 检查用户名是否重复
            if db.query(User).filter(User.username == user_data.username).first():
                error_users.append({"username": user_data.username, "error": "用户名已存在"})
                continue
            
            hashed_password = pwd_context.hash(user_data.password)
            db_user = User(
                username=user_data.username,
                hashed_password=hashed_password,
                role=user_data.role,
                real_name=user_data.real_name,
                email=user_data.email,
                student_id=user_data.student_id,
                teacher_id=user_data.teacher_id,
                class_id=user_data.class_id,
                is_active=user_data.is_active,
                created_by=current_user.id
            )
            db.add(db_user)
            success_count += 1
        except Exception as e:
            error_users.append({"username": user_data.username, "error": str(e)})
    
    db.commit()

    # 记录日志
    log_operation(db, current_user.id, 'import', None, f"批量创建用户：成功{success_count}个，失败{len(error_users)}个")

    return {"success_count": success_count, "error_users": error_users, "message": f"成功创建 {success_count} 个用户"}


# ==================== 批量操作 API ====================

class BatchDeleteRequest(BaseModel):
    user_ids: List[int]

class BatchUpdateClassRequest(BaseModel):
    user_ids: List[int]
    class_id: Optional[int] = None

class BatchResetPasswordRequest(BaseModel):
    user_ids: List[int]
    new_password: str

@app.post("/api/users/batch-delete")
async def batch_delete_users(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """批量删除用户"""
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用户")

    deleted_count = 0
    failed_users = []

    for user_id in request.user_ids:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            if user.username == 'admin':
                failed_users.append({"user_id": user_id, "username": user.username, "error": "不能删除管理员账号"})
                continue
            username = user.username
            db.delete(user)
            deleted_count += 1
            log_operation(db, current_user.id, 'delete', user_id, f"批量删除用户：{username}")

    db.commit()

    return {
        "message": f"成功删除 {deleted_count} 个用户",
        "deleted_count": deleted_count,
        "failed_users": failed_users
    }

@app.post("/api/users/batch-update-class")
async def batch_update_class(
    request: BatchUpdateClassRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """批量修改班级"""
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要修改的用户")

    # 验证班级是否存在（如果提供了 class_id）
    if request.class_id:
        cls = db.query(Class).filter(Class.id == request.class_id).first()
        if not cls:
            raise HTTPException(status_code=400, detail="班级不存在")

    updated_count = 0
    for user_id in request.user_ids:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            # 只有学生可以分配班级
            if user.role == 'student' or request.class_id is None:
                user.class_id = request.class_id
                updated_count += 1
                log_operation(db, current_user.id, 'update', user_id, f"批量修改班级：{user.username}")

    db.commit()

    return {
        "message": f"成功修改 {updated_count} 个用户的班级",
        "updated_count": updated_count
    }

@app.post("/api/users/batch-reset-password")
async def batch_reset_password(
    request: BatchResetPasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """批量重置密码"""
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要重置密码的用户")

    # 验证密码强度（至少 6 位）
    if len(request.new_password) < 6:
        raise HTTPException(status_code=400, detail="密码至少 6 位")

    reset_count = 0
    failed_users = []

    for user_id in request.user_ids:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            if user.username == 'admin':
                failed_users.append({"user_id": user_id, "username": user.username, "error": "不能重置管理员密码"})
                continue
            user.hashed_password = pwd_context.hash(request.new_password)
            reset_count += 1
            log_operation(db, current_user.id, 'reset_password', user_id, f"批量重置密码：{user.username}")

    db.commit()

    return {
        "message": f"成功重置 {reset_count} 个用户的密码",
        "reset_count": reset_count,
        "failed_users": failed_users
    }

@app.post("/api/users/import")
async def import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Excel 导入用户"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xls）")
    
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active
        
        success_count = 0
        error_rows = []
        
        # 从第 2 行开始读取（第 1 行是表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                username = row[0].value
                password = row[1].value
                role = row[2].value
                real_name = row[3].value
                id_number = row[4].value  # 学号或工号
                email = row[5].value
                class_name = row[6].value if len(row) > 6 else None

                # 验证必填字段
                if not all([username, password, role]):
                    error_rows.append({"row": row_idx, "error": "缺少必填字段（用户名、密码、角色）"})
                    continue

                # 验证用户名格式
                if not validate_username(username):
                    error_rows.append({"row": row_idx, "error": "用户名格式无效（只能包含字母、数字、下划线，长度 3-20 位）"})
                    continue
                
                # 验证密码强度
                if not validate_password(password):
                    error_rows.append({"row": row_idx, "error": "密码强度不足（必须至少 8 位，且包含大小写字母和数字）"})
                    continue
                
                # 验证邮箱格式
                if email and not validate_email(email):
                    error_rows.append({"row": row_idx, "error": "邮箱格式无效"})
                    continue

                # 验证角色
                if role not in ['student', 'teacher', 'admin']:
                    error_rows.append({"row": row_idx, "error": "无效的角色（必须是 student、teacher 或 admin）"})
                    continue
                
                # 验证学号/工号
                if role == 'student' and not id_number:
                    error_rows.append({"row": row_idx, "error": "学生必须填写学号"})
                    continue
                if role == 'teacher' and not id_number:
                    error_rows.append({"row": row_idx, "error": "教师必须填写工号"})
                    continue

                # 检查用户名是否重复
                if db.query(User).filter(User.username == username).first():
                    error_rows.append({"row": row_idx, "error": "用户名已存在"})
                    continue
                
                # 检查学号/工号是否重复
                if role == 'student' and id_number:
                    if db.query(User).filter(User.student_id == id_number).first():
                        error_rows.append({"row": row_idx, "error": "学号已存在"})
                        continue
                if role == 'teacher' and id_number:
                    if db.query(User).filter(User.teacher_id == id_number).first():
                        error_rows.append({"row": row_idx, "error": "工号已存在"})
                        continue

                # 获取班级 ID
                class_id = None
                if class_name and role == 'student':
                    cls = db.query(Class).filter(Class.class_name == class_name).first()
                    if not cls:
                        error_rows.append({"row": row_idx, "error": f"班级不存在：{class_name}"})
                        continue
                    class_id = cls.id

                # 创建用户
                hashed_password = pwd_context.hash(password)
                db_user = User(
                    username=username,
                    hashed_password=hashed_password,
                    role=role,
                    real_name=real_name,
                    email=email,
                    student_id=id_number if role == 'student' else None,
                    teacher_id=id_number if role == 'teacher' else None,
                    class_id=class_id,
                    is_active=True,
                    created_by=current_user.id
                )
                db.add(db_user)
                success_count += 1

            except Exception as e:
                error_rows.append({"row": row_idx, "error": str(e)})
        
        db.commit()
        
        # 记录日志
        log_operation(db, current_user.id, 'import', None, f"Excel 导入用户：成功{success_count}个，失败{len(error_rows)}个")
        
        return UserImportResult(
            success_count=success_count,
            error_rows=error_rows,
            message=f"成功导入 {success_count} 个用户，{len(error_rows)} 个失败"
        )
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")

@app.get("/api/users/export")
async def export_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """导出用户列表为 CSV"""
    import pandas as pd
    
    users = db.query(User).all()
    data = []
    for user in users:
        class_name = ""
        if user.class_id:
            cls = db.query(Class).filter(Class.id == user.class_id).first()
            if cls:
                class_name = cls.class_name
        
        data.append({
            "用户名": user.username,
            "角色": user.role,
            "真实姓名": user.real_name or "",
            "学号/工号": user.student_id or user.teacher_id or "",
            "邮箱": user.email or "",
            "班级": class_name,
            "状态": "启用" if user.is_active else "禁用"
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    filename = f"用户列表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    encoded_filename = urllib.parse.quote(filename)
    
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

# ---------------- 班级管理 ----------------
@app.get("/api/classes", response_model=List[ClassResponse])
async def get_classes(
    grade: Optional[str] = None,
    is_active: Optional[bool] = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取班级列表"""
    query = db.query(Class)
    
    if grade:
        query = query.filter(Class.grade == grade)
    if is_active is not None:
        query = query.filter(Class.is_active == is_active)
    
    classes = query.order_by(Class.grade, Class.class_name).all()
    
    result = []
    for cls in classes:
        teacher_name = None
        if cls.teacher_id:
            teacher = db.query(User).filter(User.id == cls.teacher_id).first()
            if teacher:
                teacher_name = teacher.real_name or teacher.username
        
        student_count = db.query(User).filter(User.class_id == cls.id).count()
        
        result.append(ClassResponse(
            id=cls.id,
            class_name=cls.class_name,
            grade=cls.grade,
            description=cls.description,
            teacher_id=cls.teacher_id,
            is_active=cls.is_active,
            created_at=cls.created_at,
            teacher_name=teacher_name,
            student_count=student_count
        ))
    
    return result

@app.post("/api/classes", response_model=ClassResponse)
async def create_class(
    cls: ClassCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """创建班级"""
    db_class = Class(**cls.dict())
    db.add(db_class)
    db.commit()
    db.refresh(db_class)
    
    return ClassResponse(
        id=db_class.id,
        class_name=db_class.class_name,
        grade=db_class.grade,
        description=db_class.description,
        teacher_id=db_class.teacher_id,
        is_active=db_class.is_active,
        created_at=db_class.created_at,
        teacher_name=None,
        student_count=0
    )

@app.put("/api/classes/{class_id}", response_model=ClassResponse)
async def update_class(
    class_id: int,
    cls_update: ClassUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """更新班级"""
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="班级不存在")
    
    update_data = cls_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_class, key, value)
    
    db.commit()
    db.refresh(db_class)
    
    teacher_name = None
    if db_class.teacher_id:
        teacher = db.query(User).filter(User.id == db_class.teacher_id).first()
        if teacher:
            teacher_name = teacher.real_name or teacher.username
    
    student_count = db.query(User).filter(User.class_id == db_class.id).count()
    
    return ClassResponse(
        id=db_class.id,
        class_name=db_class.class_name,
        grade=db_class.grade,
        description=db_class.description,
        teacher_id=db_class.teacher_id,
        is_active=db_class.is_active,
        created_at=db_class.created_at,
        teacher_name=teacher_name,
        student_count=student_count
    )

@app.delete("/api/classes/{class_id}")
async def delete_class(class_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    """删除班级"""
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="班级不存在")
    
    # 检查是否有学生
    has_students = db.query(User).filter(User.class_id == class_id).first()
    if has_students:
        raise HTTPException(status_code=400, detail="班级下有学生，无法删除")
    
    db.delete(db_class)
    db.commit()
    
    return {"message": "班级已删除"}

@app.get("/api/classes/{class_id}/students", response_model=List[UserResponse])
async def get_class_students(class_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """获取班级学生列表"""
    cls = db.query(Class).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")
    
    students = db.query(User).filter(User.class_id == class_id, User.role == 'student').all()
    
    result = []
    for student in students:
        result.append(UserResponse(
            id=student.id,
            username=student.username,
            role=student.role,
            email=student.email,
            real_name=student.real_name,
            student_id=student.student_id,
            teacher_id=student.teacher_id,
            class_id=student.class_id,
            is_active=student.is_active,
            created_at=student.created_at,
            updated_at=student.updated_at,
            class_name=cls.class_name
        ))
    
    return result


# ==================== 实验报告系统 API ====================

@app.get("/api/assignments", response_model=List[AssignmentResponse])
async def get_assignments(
    class_id: Optional[int] = None,
    teacher_id: Optional[int] = None,
    is_published: Optional[bool] = None,
    status: Optional[str] = None,  # 仅学生使用：all, pending, submitted, graded
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取实验任务列表 - 已优化 N+1 查询"""
    query = db.query(Assignment)

    if class_id:
        query = query.filter(Assignment.class_id == class_id)
    if teacher_id:
        query = query.filter(Assignment.teacher_id == teacher_id)
    if is_published is not None:
        query = query.filter(Assignment.is_published == is_published)
    else:
        # 默认只显示已发布的任务（学生）
        query = query.filter(Assignment.is_published == True)

    # 使用 selectinload 预加载关联数据，避免 N+1 查询
    assignments = query.options(
        joinedload(Assignment.teacher),
        joinedload(Assignment.device),
        joinedload(Assignment.clas)
    ).order_by(desc(Assignment.created_at)).all()

    # 一次性获取所有提交数量
    submission_counts = db.query(
        AssignmentSubmission.assignment_id,
        func.count(AssignmentSubmission.id).label('count')
    ).group_by(AssignmentSubmission.assignment_id).all()
    count_map = {sc.assignment_id: sc.count for sc in submission_counts}

    # 构建响应
    result = []
    for assignment in assignments:
        result.append(AssignmentResponse(
            id=assignment.id,
            title=assignment.title,
            description=assignment.description,
            device_id=assignment.device_id,
            class_id=assignment.class_id,
            teacher_id=assignment.teacher_id,
            teacher_name=assignment.teacher.real_name or assignment.teacher.username if assignment.teacher else None,
            device_name=assignment.device.device_name if assignment.device else None,
            class_name=assignment.clas.class_name if assignment.clas else None,
            start_date=assignment.start_date,
            due_date=assignment.due_date,
            requirement=assignment.requirement,
            template=assignment.template,
            is_published=assignment.is_published,
            created_at=assignment.created_at,
            submission_count=count_map.get(assignment.id, 0)
        ))

    return result


@app.get("/api/assignments/{assignment_id}", response_model=AssignmentResponse)
async def get_assignment(assignment_id: int, db: Session = Depends(get_db)):
    """获取实验任务详情"""
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="任务不存在")

    return assignment


@app.post("/api/assignments", response_model=AssignmentResponse)
async def create_assignment(
    assignment: AssignmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """创建实验任务（教师/管理员）"""
    db_assignment = Assignment(
        **assignment.dict(),
        teacher_id=current_user.id
    )
    db.add(db_assignment)
    db.commit()
    db.refresh(db_assignment)
    return db_assignment


@app.put("/api/assignments/{assignment_id}", response_model=AssignmentResponse)
async def update_assignment(
    assignment_id: int,
    assignment_update: AssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """更新实验任务（教师/管理员）"""
    db_assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not db_assignment:
        raise HTTPException(status_code=404, detail="任务不存在")

    # 教师只能编辑自己创建的任务
    if current_user.role != 'admin' and db_assignment.teacher_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能编辑自己创建的任务")

    update_data = assignment_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_assignment, key, value)

    db.commit()
    db.refresh(db_assignment)
    return db_assignment


@app.delete("/api/assignments/{assignment_id}")
async def delete_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """删除实验任务（教师/管理员）"""
    db_assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not db_assignment:
        raise HTTPException(status_code=404, detail="任务不存在")

    if current_user.role != 'admin' and db_assignment.teacher_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能删除自己创建的任务")

    db.delete(db_assignment)
    db.commit()
    return {"message": "任务已删除"}


@app.get("/api/assignments/{assignment_id}/submissions", response_model=List[AssignmentSubmissionResponse])
async def get_submissions(
    assignment_id: int,
    student_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取实验报告提交列表"""
    query = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id
    )

    if student_id:
        query = query.filter(AssignmentSubmission.student_id == student_id)
    if status:
        query = query.filter(AssignmentSubmission.status == status)

    submissions = query.order_by(desc(AssignmentSubmission.created_at)).all()

    # 构建响应
    result = []
    for sub in submissions:
        student_name = None
        if sub.student_id:
            student = db.query(User).filter(User.id == sub.student_id).first()
            if student:
                student_name = student.username

        result.append(AssignmentSubmissionResponse(
            id=sub.id,
            assignment_id=sub.assignment_id,
            student_id=sub.student_id,
            student_name=student_name,
            status=sub.status,
            experiment_date=sub.experiment_date,
            observations=sub.observations,
            conclusion=sub.conclusion,
            temp_records=sub.temp_records,
            humidity_records=sub.humidity_records,
            soil_moisture_records=sub.soil_moisture_records,
            light_records=sub.light_records,
            photos=sub.photos,
            score=sub.score,
            teacher_comment=sub.teacher_comment,
            graded_at=sub.graded_at,
            submitted_at=sub.submitted_at,
            created_at=sub.created_at,
            updated_at=sub.updated_at
        ))

    return result


@app.get("/api/assignments/{assignment_id}/my-submission", response_model=AssignmentSubmissionResponse)
async def get_my_submission(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取我的实验报告（学生）"""
    submission = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id
    ).first()

    if not submission:
        raise HTTPException(status_code=404, detail="未找到提交记录")

    return submission


@app.post("/api/assignments/{assignment_id}/submit", response_model=AssignmentSubmissionResponse)
async def submit_assignment(
    assignment_id: int,
    submission: AssignmentSubmissionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """提交/更新实验报告（学生）"""
    # 检查任务是否存在
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    # 验证截止时间
    if assignment.due_date and datetime.datetime.utcnow() > assignment.due_date:
        raise HTTPException(
            status_code=400, 
            detail=f"已超过提交截止时间 ({assignment.due_date.strftime('%Y-%m-%d %H:%M')})"
        )
    
    # 验证学生是否属于任务布置的班级（如果任务有班级限制）
    if assignment.class_id:
        student = db.query(User).filter(User.id == current_user.id).first()
        if student and student.class_id != assignment.class_id and current_user.role != 'teacher':
            raise HTTPException(status_code=403, detail="您不属于该任务布置的班级")

    # 查找或创建提交记录
    existing = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id
    ).first()

    if existing:
        # 更新现有记录
        existing.observations = submission.observations
        existing.conclusion = submission.conclusion
        existing.temp_records = submission.temp_records
        existing.humidity_records = submission.humidity_records
        existing.soil_moisture_records = submission.soil_moisture_records
        existing.light_records = submission.light_records
        existing.photos = submission.photos
        existing.experiment_date = submission.experiment_date
        existing.status = "submitted"
        existing.submitted_at = datetime.datetime.utcnow()
        db.refresh(existing)
        result = existing
    else:
        # 创建新记录
        new_submission = AssignmentSubmission(
            assignment_id=assignment_id,
            student_id=current_user.id,
            experiment_date=submission.experiment_date,
            observations=submission.observations,
            conclusion=submission.conclusion,
            temp_records=submission.temp_records,
            humidity_records=submission.humidity_records,
            soil_moisture_records=submission.soil_moisture_records,
            light_records=submission.light_records,
            photos=submission.photos,
            status="submitted",
            submitted_at=datetime.datetime.utcnow()
        )
        db.add(new_submission)
        db.commit()
        db.refresh(new_submission)
        result = new_submission

    return result


@app.post("/api/assignments/{assignment_id}/grade")
async def grade_assignment(
    assignment_id: int,
    submission_id: int,
    grade: AssignmentSubmissionGrade,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """批改实验报告（教师）- 已修复权限验证"""
    submission = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.id == submission_id,
        AssignmentSubmission.assignment_id == assignment_id
    ).first()
    if not submission:
        raise HTTPException(status_code=404, detail="提交记录不存在")
    
    # 验证批改权限：只有任务创建者或管理员可以批改
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if current_user.role != 'admin' and assignment.teacher_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能批改自己布置的任务")

    submission.score = grade.score
    submission.teacher_comment = grade.teacher_comment
    submission.graded_at = datetime.datetime.utcnow()
    submission.graded_by = current_user.id
    submission.status = "graded"

    db.commit()
    return {"message": "批改完成"}


# ==================== 植物生长档案 API ====================

@app.get("/api/plants", response_model=List[PlantProfileResponse])
async def get_plants(
    class_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取植物档案列表 - 已优化 N+1 查询"""
    query = db.query(PlantProfile)

    if class_id:
        query = query.filter(PlantProfile.class_id == class_id)
    if status:
        query = query.filter(PlantProfile.status == status)

    # 使用 selectinload 预加载关联数据
    plants = query.options(
        joinedload(PlantProfile.clas),
        joinedload(PlantProfile.device),
        joinedload(PlantProfile.group)
    ).order_by(desc(PlantProfile.created_at)).all()

    # 构建响应
    result = []
    for plant in plants:
        # 获取生长记录数量（使用关联对象）
        record_count = len(plant.growth_records) if hasattr(plant, 'growth_records') else 0

        result.append(PlantProfileResponse(
            id=plant.id,
            plant_name=plant.plant_name,
            species=plant.species,
            class_id=plant.class_id,
            group_id=plant.group_id,
            device_id=plant.device_id,
            class_name=plant.clas.class_name if plant.clas else None,
            device_name=plant.device.device_name if plant.device else None,
            group_name=plant.group.group_name if plant.group else None,
            plant_date=plant.plant_date,
            cover_image=plant.cover_image,
            status=plant.status,
            expected_harvest_date=plant.expected_harvest_date,
            description=plant.description,
            growth_record_count=record_count,
            created_at=plant.created_at
        ))

    return result


@app.get("/api/plants/{plant_id}", response_model=PlantProfileResponse)
async def get_plant(plant_id: int, db: Session = Depends(get_db)):
    """获取植物档案详情"""
    plant = db.query(PlantProfile).filter(PlantProfile.id == plant_id).first()
    if not plant:
        raise HTTPException(status_code=404, detail="植物档案不存在")
    return plant


@app.post("/api/plants", response_model=PlantProfileResponse)
async def create_plant(
    plant: PlantProfileCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """创建植物档案（教师/管理员）"""
    db_plant = PlantProfile(**plant.dict())
    db.add(db_plant)
    db.commit()
    db.refresh(db_plant)
    return db_plant


@app.put("/api/plants/{plant_id}", response_model=PlantProfileResponse)
async def update_plant(
    plant_id: int,
    plant_update: PlantProfileUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """更新植物档案（教师/管理员）"""
    db_plant = db.query(PlantProfile).filter(PlantProfile.id == plant_id).first()
    if not db_plant:
        raise HTTPException(status_code=404, detail="植物档案不存在")

    update_data = plant_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_plant, key, value)

    db.commit()
    db.refresh(db_plant)
    return db_plant


@app.delete("/api/plants/{plant_id}")
async def delete_plant(
    plant_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """删除植物档案（教师/管理员）"""
    db_plant = db.query(PlantProfile).filter(PlantProfile.id == plant_id).first()
    if not db_plant:
        raise HTTPException(status_code=404, detail="植物档案不存在")

    db.delete(db_plant)
    db.commit()
    return {"message": "植物档案已删除"}


@app.get("/api/plants/{plant_id}/records", response_model=List[GrowthRecordResponse])
async def get_plant_records(plant_id: int, db: Session = Depends(get_db)):
    """获取植物生长记录列表"""
    records = db.query(GrowthRecord).filter(
        GrowthRecord.plant_id == plant_id
    ).order_by(desc(GrowthRecord.record_date)).all()

    # 构建响应
    result = []
    for record in records:
        recorder_name = None
        if record.recorded_by:
            recorder = db.query(User).filter(User.id == record.recorded_by).first()
            if recorder:
                recorder_name = recorder.username

        result.append(GrowthRecordResponse(
            id=record.id,
            plant_id=record.plant_id,
            record_date=record.record_date,
            stage=record.stage,
            height_cm=float(record.height_cm) if record.height_cm else None,
            leaf_count=record.leaf_count,
            flower_count=record.flower_count,
            fruit_count=record.fruit_count,
            description=record.description,
            photos=record.photos,
            recorder_name=recorder_name,
            created_at=record.created_at
        ))

    return result


@app.post("/api/plants/{plant_id}/records", response_model=GrowthRecordResponse)
async def create_plant_record(
    plant_id: int,
    record: GrowthRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """添加生长记录（学生/教师）"""
    plant = db.query(PlantProfile).filter(PlantProfile.id == plant_id).first()
    if not plant:
        raise HTTPException(status_code=404, detail="植物档案不存在")
    
    # 验证权限：只有教师或小组成员可以添加记录
    if current_user.role not in ['teacher', 'admin']:
        # 检查学生是否属于负责该植物的小组
        if plant.group_id:
            member = db.query(GroupMember).filter(
                GroupMember.group_id == plant.group_id,
                GroupMember.student_id == current_user.id
            ).first()
            if not member:
                raise HTTPException(status_code=403, detail="无权为该植物添加记录")
        elif plant.class_id:
            # 如果没有小组，检查是否属于班级
            student = db.query(User).filter(User.id == current_user.id).first()
            if not student or student.class_id != plant.class_id:
                raise HTTPException(status_code=403, detail="无权为该植物添加记录")
        else:
            raise HTTPException(status_code=403, detail="无权为该植物添加记录")

    db_record = GrowthRecord(
        **record.dict(),
        recorded_by=current_user.id
    )
    db.add(db_record)
    db.commit()
    db.refresh(db_record)

    recorder_name = current_user.username

    return GrowthRecordResponse(
        id=db_record.id,
        plant_id=db_record.plant_id,
        record_date=db_record.record_date,
        stage=db_record.stage,
        height_cm=float(db_record.height_cm) if db_record.height_cm else None,
        leaf_count=db_record.leaf_count,
        flower_count=db_record.flower_count,
        fruit_count=db_record.fruit_count,
        description=db_record.description,
        photos=db_record.photos,
        recorder_name=recorder_name,
        created_at=db_record.created_at
    )


@app.delete("/api/plants/records/{record_id}")
async def delete_plant_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """删除生长记录（教师/管理员）"""
    db_record = db.query(GrowthRecord).filter(GrowthRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")

    db.delete(db_record)
    db.commit()
    return {"message": "记录已删除"}


# ==================== 小组合作学习 API ====================

@app.get("/api/groups", response_model=List[StudyGroupResponse])
async def get_groups(
    class_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取小组列表"""
    query = db.query(StudyGroup)

    if class_id:
        query = query.filter(StudyGroup.class_id == class_id)

    groups = query.order_by(desc(StudyGroup.created_at)).all()

    # 构建响应
    result = []
    for group in groups:
        # 获取成员
        members = db.query(GroupMember).filter(GroupMember.group_id == group.id).all()
        member_list = []
        for member in members:
            student = db.query(User).filter(User.id == member.student_id).first()
            member_list.append({
                "id": member.id,
                "student_id": member.student_id,
                "student_name": student.username if student else None,
                "role": member.role
            })

        # 获取班级名称
        class_name = None
        if group.class_id:
            cls = db.query(Class).filter(Class.id == group.class_id).first()
            if cls:
                class_name = cls.class_name

        # 获取设备名称
        device_name = None
        if group.device_id:
            device = db.query(Device).filter(Device.id == group.device_id).first()
            if device:
                device_name = device.device_name

        result.append(StudyGroupResponse(
            id=group.id,
            group_name=group.group_name,
            class_id=group.class_id,
            class_name=class_name,
            device_id=group.device_id,
            device_name=device_name,
            description=group.description,
            member_count=len(members),
            members=member_list,
            created_at=group.created_at
        ))

    return result


@app.post("/api/groups", response_model=StudyGroupResponse)
async def create_group(
    group: StudyGroupCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """创建小组（教师/管理员）"""
    db_group = StudyGroup(**group.dict())
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return db_group


@app.get("/api/groups/{group_id}", response_model=StudyGroupResponse)
async def get_group(group_id: int, db: Session = Depends(get_db)):
    """获取小组详情"""
    group = db.query(StudyGroup).filter(StudyGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="小组不存在")
    return group


@app.post("/api/groups/{group_id}/members", response_model=GroupMemberResponse)
async def add_group_member(
    group_id: int,
    member: GroupMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """添加小组成员（教师/管理员）"""
    group = db.query(StudyGroup).filter(StudyGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="小组不存在")

    db_member = GroupMember(**member.dict())
    db.add(db_member)
    db.commit()
    db.refresh(db_member)

    student_name = None
    if member.student_id:
        student = db.query(User).filter(User.id == member.student_id).first()
        if student:
            student_name = student.username

    return GroupMemberResponse(
        id=db_member.id,
        group_id=db_member.group_id,
        student_id=db_member.student_id,
        student_name=student_name,
        role=db_member.role,
        joined_at=db_member.joined_at
    )


@app.delete("/api/groups/members/{member_id}")
async def remove_group_member(
    member_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_teacher_user)
):
    """移除小组成员（教师/管理员）"""
    db_member = db.query(GroupMember).filter(GroupMember.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="成员不存在")

    db.delete(db_member)
    db.commit()
    return {"message": "成员已移除"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
