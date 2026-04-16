import datetime
import csv
import io
import urllib.parse

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.api.dependencies import get_admin_user, get_current_user, get_db
from app.core.security import hash_password
from app.core.validators import (
    get_password_rule_text,
    is_strict_password_role,
    validate_email,
    validate_password,
    validate_username,
)
from app.db.models import Class, ClassDeviceBind, Device, User, UserOperationLog
from app.db.models import (
    Assignment,
    AssignmentSubmission,
    ContentComment,
    GrowthRecord,
    GroupMember,
    StudentLearningRecord,
    TeachingContent,
)
from app.schemas.users import (
    BatchDeleteRequest,
    BatchResetPasswordRequest,
    BatchUpdateClassRequest,
    ClassCreate,
    ClassDeviceBindCreate,
    ClassResponse,
    ClassUpdate,
    UserCreate,
    UserImportResult,
    UserResponse,
    UserStats,
    UserUpdate,
)


router = APIRouter(tags=["users"])


def log_operation(
    db: Session,
    operator_id: int,
    operation_type: str,
    target_user_id: int | None = None,
    details: str | None = None,
):
    log = UserOperationLog(
        operator_id=operator_id,
        operation_type=operation_type,
        target_user_id=target_user_id,
        details=details,
    )
    db.add(log)


@router.get("/api/users")
async def get_users(
    role: str | None = None,
    class_id: int | None = None,
    is_active: bool | None = None,
    search: str | None = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    query = db.query(User)

    if role:
        query = query.filter(User.role == role)
    if class_id:
        query = query.filter(User.class_id == class_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (User.username.like(search_pattern))
            | (User.real_name.like(search_pattern))
            | (User.student_id.like(search_pattern))
        )

    total = query.count()
    users = (
        query.order_by(desc(User.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    class_map = {}
    class_ids = {u.class_id for u in users if u.class_id}
    if class_ids:
        classes = db.query(Class).filter(Class.id.in_(class_ids)).all()
        class_map = {c.id: c.class_name for c in classes}

    result = []
    for user in users:
        class_name = class_map.get(user.class_id)
        result.append(
            {
                "id": user.id,
                "username": user.username,
                "role": user.role,
                "email": user.email,
                "real_name": user.real_name,
                "student_id": user.student_id,
                "teacher_id": user.teacher_id,
                "class_id": user.class_id,
                "is_active": user.is_active,
                "created_at": user.created_at,
                "updated_at": user.updated_at,
                "class_name": class_name,
            }
        )

    return {"items": result, "total": total, "page": page, "page_size": page_size}


@router.post("/api/users", response_model=UserResponse)
async def create_user(
    user: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    if not validate_username(user.username):
        raise HTTPException(
            status_code=400,
            detail="用户名格式无效：只能包含字母、数字、下划线，长度 3-20 位",
        )

    strict_mode = is_strict_password_role(user.role)
    password_rule = get_password_rule_text(strict_mode)
    if not validate_password(user.password, strict=strict_mode):
        raise HTTPException(status_code=400, detail=f"密码强度不足：{password_rule}")

    if user.email and not validate_email(user.email):
        raise HTTPException(status_code=400, detail="邮箱格式无效")

    existing = db.query(User).filter(User.username == user.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")

    if user.student_id:
        if db.query(User).filter(User.student_id == user.student_id).first():
            raise HTTPException(status_code=400, detail="学号已存在")
    if user.teacher_id:
        if db.query(User).filter(User.teacher_id == user.teacher_id).first():
            raise HTTPException(status_code=400, detail="工号已存在")

    hashed_password = hash_password(user.password)
    db_user = User(
        username=user.username,
        hashed_password=hashed_password,
        role=user.role,
        real_name=user.real_name,
        email=user.email,
        student_id=user.student_id,
        teacher_id=user.teacher_id,
        class_id=user.class_id,
        is_active=user.is_active,
        created_by=current_user.id,
    )
    db.add(db_user)
    db.flush()
    log_operation(db, current_user.id, "create_user", db_user.id, f"创建用户：{user.username} ({user.role})")
    db.commit()
    db.refresh(db_user)

    class_name = None
    if db_user.class_id:
        cls = db.query(Class).filter(Class.id == db_user.class_id).first()
        if cls:
            class_name = cls.class_name

    return UserResponse(
        id=db_user.id,
        username=db_user.username,
        role=db_user.role,
        email=db_user.email,
        real_name=db_user.real_name,
        student_id=db_user.student_id,
        teacher_id=db_user.teacher_id,
        class_id=db_user.class_id,
        is_active=db_user.is_active,
        created_at=db_user.created_at,
        updated_at=db_user.updated_at,
        class_name=class_name,
    )


@router.get("/api/users/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    class_name = None
    if user.class_id:
        cls = db.query(Class).filter(Class.id == user.class_id).first()
        if cls:
            class_name = cls.class_name

    return UserResponse(
        id=user.id,
        username=user.username,
        role=user.role,
        email=user.email,
        real_name=user.real_name,
        student_id=user.student_id,
        teacher_id=user.teacher_id,
        class_id=user.class_id,
        is_active=user.is_active,
        created_at=user.created_at,
        updated_at=user.updated_at,
        class_name=class_name,
    )


@router.put("/api/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    update_data = user_update.model_dump(exclude_unset=True)

    new_email = update_data.get("email")
    if new_email and not validate_email(new_email):
        raise HTTPException(status_code=400, detail="邮箱格式无效")

    new_role = update_data.get("role")
    if new_role and new_role not in ["student", "teacher", "admin"]:
        raise HTTPException(status_code=400, detail="角色无效")

    new_student_id = update_data.get("student_id")
    if new_student_id:
        student_exists = (
            db.query(User)
            .filter(User.student_id == new_student_id, User.id != user_id)
            .first()
        )
        if student_exists:
            raise HTTPException(status_code=400, detail="学号已存在")

    new_teacher_id = update_data.get("teacher_id")
    if new_teacher_id:
        teacher_exists = (
            db.query(User)
            .filter(User.teacher_id == new_teacher_id, User.id != user_id)
            .first()
        )
        if teacher_exists:
            raise HTTPException(status_code=400, detail="工号已存在")

    for key, value in update_data.items():
        setattr(db_user, key, value)

    log_operation(db, current_user.id, "update_user", db_user.id, f"更新用户：{db_user.username}")
    db.commit()
    db.refresh(db_user)

    class_name = None
    if db_user.class_id:
        cls = db.query(Class).filter(Class.id == db_user.class_id).first()
        if cls:
            class_name = cls.class_name

    return UserResponse(
        id=db_user.id,
        username=db_user.username,
        role=db_user.role,
        email=db_user.email,
        real_name=db_user.real_name,
        student_id=db_user.student_id,
        teacher_id=db_user.teacher_id,
        class_id=db_user.class_id,
        is_active=db_user.is_active,
        created_at=db_user.created_at,
        updated_at=db_user.updated_at,
        class_name=class_name,
    )


@router.delete("/api/users/{user_id}")
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能删除当前登录账号")

    if db_user.role == "admin" or db_user.username == "admin":
        raise HTTPException(status_code=400, detail="管理员账号不允许删除")

    dependency_checks = [
        (db.query(Class).filter(Class.teacher_id == user_id).count(), "班级班主任"),
        (db.query(Assignment).filter(Assignment.teacher_id == user_id).count(), "实验任务创建者"),
        (db.query(AssignmentSubmission).filter(AssignmentSubmission.student_id == user_id).count(), "实验报告提交"),
        (db.query(AssignmentSubmission).filter(AssignmentSubmission.graded_by == user_id).count(), "实验报告批改"),
        (db.query(TeachingContent).filter(TeachingContent.author_id == user_id).count(), "教学内容作者"),
        (db.query(StudentLearningRecord).filter(StudentLearningRecord.student_id == user_id).count(), "学习记录"),
        (db.query(ContentComment).filter(ContentComment.student_id == user_id).count(), "内容评论"),
        (db.query(GrowthRecord).filter(GrowthRecord.recorded_by == user_id).count(), "植物记录"),
        (db.query(GroupMember).filter(GroupMember.student_id == user_id).count(), "小组成员"),
        (db.query(UserOperationLog).filter(UserOperationLog.operator_id == user_id).count(), "操作日志(操作员)"),
    ]
    blocking = [name for count, name in dependency_checks if count > 0]
    if blocking:
        tips = "、".join(blocking[:4])
        raise HTTPException(status_code=400, detail=f"用户存在关联数据（{tips}），请先处理关联数据或改用禁用")

    username = db_user.username
    # 目标用户引用为可空字段，删除前置空可保留日志正文信息。
    db.query(UserOperationLog).filter(UserOperationLog.target_user_id == user_id).update(
        {UserOperationLog.target_user_id: None}, synchronize_session=False
    )
    db.delete(db_user)
    log_operation(db, current_user.id, "delete_user", None, f"删除用户：{username}(id={user_id})")
    db.commit()

    return {"message": "用户已删除"}


@router.post("/api/users/{user_id}/reset-password")
async def reset_password(
    user_id: int,
    new_password: str = Query(..., min_length=6),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    strict_mode = is_strict_password_role(db_user.role)
    password_rule = get_password_rule_text(strict_mode)

    if not validate_password(new_password, strict=strict_mode):
        raise HTTPException(status_code=400, detail=f"密码强度不足：{password_rule}")

    db_user.hashed_password = hash_password(new_password)
    log_operation(db, current_user.id, "reset_password", user_id, f"重置用户密码：{db_user.username}")
    db.commit()

    return {"message": "密码已重置"}


@router.post("/api/users/{user_id}/toggle-active")
async def toggle_user_active(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    if (db_user.role == "admin" or db_user.username == "admin") and db_user.is_active:
        raise HTTPException(status_code=400, detail="管理员账号不允许禁用")

    db_user.is_active = not db_user.is_active
    status_text = "启用" if db_user.is_active else "禁用"
    log_operation(db, current_user.id, "toggle_active", user_id, f"{status_text}用户：{db_user.username}")
    db.commit()

    return {"message": f"账号已{status_text}", "is_active": db_user.is_active}


@router.get("/api/stats/users", response_model=UserStats)
async def get_user_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    total = db.query(User).count()
    admin_count = db.query(User).filter(User.role == "admin").count()
    teacher_count = db.query(User).filter(User.role == "teacher").count()
    student_count = db.query(User).filter(User.role == "student").count()
    active = db.query(User).filter(User.is_active == True).count()
    inactive = db.query(User).filter(User.is_active == False).count()

    return UserStats(
        total_users=total,
        admin_count=admin_count,
        teacher_count=teacher_count,
        student_count=student_count,
        active_count=active,
        inactive_count=inactive,
    )


@router.post("/api/users/batch-create")
async def batch_create_users(
    users: list[UserCreate],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    success_count = 0
    error_users = []

    for user_data in users:
        try:
            if not validate_username(user_data.username):
                error_users.append({"username": user_data.username, "error": "用户名格式无效"})
                continue

            strict_mode = is_strict_password_role(user_data.role)
            if not validate_password(user_data.password, strict=strict_mode):
                error_users.append(
                    {
                        "username": user_data.username,
                        "error": f"密码强度不足：{get_password_rule_text(strict_mode)}",
                    }
                )
                continue

            if user_data.email and not validate_email(user_data.email):
                error_users.append({"username": user_data.username, "error": "邮箱格式无效"})
                continue

            if db.query(User).filter(User.username == user_data.username).first():
                error_users.append({"username": user_data.username, "error": "用户名已存在"})
                continue

            if user_data.student_id and db.query(User).filter(User.student_id == user_data.student_id).first():
                error_users.append({"username": user_data.username, "error": "学号已存在"})
                continue

            if user_data.teacher_id and db.query(User).filter(User.teacher_id == user_data.teacher_id).first():
                error_users.append({"username": user_data.username, "error": "工号已存在"})
                continue

            hashed_password = hash_password(user_data.password)
            db_user = User(
                username=user_data.username,
                hashed_password=hashed_password,
                role=user_data.role,
                real_name=user_data.real_name,
                email=user_data.email,
                student_id=user_data.student_id,
                teacher_id=user_data.teacher_id,
                class_id=user_data.class_id,
                is_active=user_data.is_active,
                created_by=current_user.id,
            )
            db.add(db_user)
            success_count += 1
        except Exception as e:
            error_users.append({"username": user_data.username, "error": str(e)})

    log_operation(
        db,
        current_user.id,
        "import_users",
        None,
        f"批量创建用户：成功{success_count}个，失败{len(error_users)}个",
    )
    db.commit()

    return {
        "success_count": success_count,
        "error_users": error_users,
        "message": f"成功创建 {success_count} 个用户",
    }


@router.post("/api/users/batch-delete")
async def batch_delete_users(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用户")

    deleted_count = 0
    failed_users = []

    for user_id in request.user_ids:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            failed_users.append({"user_id": user_id, "username": None, "error": "用户不存在"})
            continue

        if user.role == "admin" or user.username == "admin":
            failed_users.append({"user_id": user_id, "username": user.username, "error": "不能删除管理员账号"})
            continue

        if not user.is_active:
            failed_users.append({"user_id": user_id, "username": user.username, "error": "用户已禁用"})
            continue

        username = user.username
        user.is_active = False
        deleted_count += 1
        log_operation(db, current_user.id, "batch_delete", user_id, f"批量禁用用户：{username}")

    db.commit()

    return {
        "message": f"成功禁用 {deleted_count} 个用户",
        "deleted_count": deleted_count,
        "failed_users": failed_users,
    }


@router.post("/api/users/batch-update-class")
async def batch_update_class(
    request: BatchUpdateClassRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要修改的用户")

    if request.class_id:
        cls = db.query(Class).filter(Class.id == request.class_id).first()
        if not cls:
            raise HTTPException(status_code=400, detail="班级不存在")

    updated_count = 0
    for user_id in request.user_ids:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            if user.role == "student" or request.class_id is None:
                user.class_id = request.class_id
                updated_count += 1
                log_operation(db, current_user.id, "batch_update_class", user_id, f"批量修改班级：{user.username}")

    db.commit()

    return {
        "message": f"成功修改 {updated_count} 个用户的班级",
        "updated_count": updated_count,
    }


@router.post("/api/users/batch-reset-password")
async def batch_reset_password(
    request: BatchResetPasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    if not request.user_ids:
        raise HTTPException(status_code=400, detail="请选择要重置密码的用户")

    users_to_reset = db.query(User).filter(User.id.in_(request.user_ids)).all()
    user_map = {u.id: u for u in users_to_reset}

    strict_mode = any(
        user.role in ["admin", "teacher"] and user.username != "admin" for user in users_to_reset
    )
    if not validate_password(request.new_password, strict=strict_mode):
        raise HTTPException(status_code=400, detail=f"密码强度不足：{get_password_rule_text(strict_mode)}")

    reset_count = 0
    failed_users = []

    for user_id in request.user_ids:
        user = user_map.get(user_id)
        if user:
            if user.username == "admin":
                failed_users.append({"user_id": user_id, "username": user.username, "error": "不能重置管理员密码"})
                continue
            user.hashed_password = hash_password(request.new_password)
            reset_count += 1
            log_operation(db, current_user.id, "batch_reset_password", user_id, f"批量重置密码：{user.username}")
        else:
            failed_users.append({"user_id": user_id, "username": None, "error": "用户不存在"})

    db.commit()

    return {
        "message": f"成功重置 {reset_count} 个用户的密码",
        "reset_count": reset_count,
        "failed_users": failed_users,
    }


@router.post("/api/users/import")
async def import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xls）")

    try:
        from io import BytesIO

        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active

        success_count = 0
        error_rows = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                username = row[0].value
                password = row[1].value
                role = row[2].value
                real_name = row[3].value
                id_number = row[4].value
                email = row[5].value
                class_name = row[6].value if len(row) > 6 else None

                if not all([username, password, role]):
                    error_rows.append({"row": row_idx, "error": "缺少必填字段（用户名、密码、角色）"})
                    continue

                if not validate_username(username):
                    error_rows.append(
                        {"row": row_idx, "error": "用户名格式无效（只能包含字母、数字、下划线，长度 3-20 位）"}
                    )
                    continue

                strict_mode = is_strict_password_role(role)
                if not validate_password(password, strict=strict_mode):
                    error_rows.append({"row": row_idx, "error": f"密码强度不足（{get_password_rule_text(strict_mode)}）"})
                    continue

                if email and not validate_email(email):
                    error_rows.append({"row": row_idx, "error": "邮箱格式无效"})
                    continue

                if role not in ["student", "teacher", "admin"]:
                    error_rows.append({"row": row_idx, "error": "无效的角色（必须是 student、teacher 或 admin）"})
                    continue

                if role == "student" and not id_number:
                    error_rows.append({"row": row_idx, "error": "学生必须填写学号"})
                    continue
                if role == "teacher" and not id_number:
                    error_rows.append({"row": row_idx, "error": "教师必须填写工号"})
                    continue

                if db.query(User).filter(User.username == username).first():
                    error_rows.append({"row": row_idx, "error": "用户名已存在"})
                    continue

                if role == "student" and id_number:
                    if db.query(User).filter(User.student_id == id_number).first():
                        error_rows.append({"row": row_idx, "error": "学号已存在"})
                        continue
                if role == "teacher" and id_number:
                    if db.query(User).filter(User.teacher_id == id_number).first():
                        error_rows.append({"row": row_idx, "error": "工号已存在"})
                        continue

                class_id = None
                if class_name and role == "student":
                    cls = db.query(Class).filter(Class.class_name == class_name).first()
                    if not cls:
                        error_rows.append({"row": row_idx, "error": f"班级不存在：{class_name}"})
                        continue
                    class_id = cls.id

                hashed_password = hash_password(password)
                db_user = User(
                    username=username,
                    hashed_password=hashed_password,
                    role=role,
                    real_name=real_name,
                    email=email,
                    student_id=id_number if role == "student" else None,
                    teacher_id=id_number if role == "teacher" else None,
                    class_id=class_id,
                    is_active=True,
                    created_by=current_user.id,
                )
                db.add(db_user)
                success_count += 1

            except Exception as e:
                error_rows.append({"row": row_idx, "error": str(e)})

        if error_rows:
            db.rollback()
            return UserImportResult(
                success_count=0,
                error_rows=error_rows,
                message=f"导入失败，发现 {len(error_rows)} 处错误，已回滚全部操作",
            )

        log_operation(db, current_user.id, "import_users", None, f"Excel 导入用户：成功{success_count}个")
        db.commit()

        return UserImportResult(
            success_count=success_count,
            error_rows=[],
            message=f"成功导入 {success_count} 个用户，0 个失败",
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.get("/api/users-export")
async def export_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    class_map = {c.id: c.class_name for c in db.query(Class).all()}

    def iter_csv():
        stream = io.StringIO()
        writer = csv.writer(stream)
        writer.writerow(["用户名", "角色", "真实姓名", "学号/工号", "邮箱", "班级", "状态"])
        yield "\ufeff" + stream.getvalue()
        stream.seek(0)
        stream.truncate(0)

        users_query = db.query(User).order_by(User.id).yield_per(500)
        for user in users_query:
            writer.writerow(
                [
                    user.username,
                    user.role,
                    user.real_name or "",
                    user.student_id or user.teacher_id or "",
                    user.email or "",
                    class_map.get(user.class_id, "") if user.class_id else "",
                    "启用" if user.is_active else "禁用",
                ]
            )
            yield stream.getvalue()
            stream.seek(0)
            stream.truncate(0)

    filename = f"用户列表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    encoded_filename = urllib.parse.quote(filename)

    return StreamingResponse(
        iter_csv(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/api/classes", response_model=list[ClassResponse])
async def get_classes(
    grade: str | None = None,
    is_active: bool | None = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Class)

    if grade:
        query = query.filter(Class.grade == grade)
    if is_active is not None:
        query = query.filter(Class.is_active == is_active)

    classes = query.order_by(Class.grade, Class.class_name).all()

    result = []
    for cls in classes:
        teacher_name = None
        if cls.teacher_id:
            teacher = db.query(User).filter(User.id == cls.teacher_id).first()
            if teacher:
                teacher_name = teacher.real_name or teacher.username

        student_count = db.query(User).filter(User.class_id == cls.id).count()

        result.append(
            ClassResponse(
                id=cls.id,
                class_name=cls.class_name,
                grade=cls.grade,
                description=cls.description,
                teacher_id=cls.teacher_id,
                is_active=cls.is_active,
                created_at=cls.created_at,
                teacher_name=teacher_name,
                student_count=student_count,
            )
        )

    return result


@router.post("/api/classes", response_model=ClassResponse)
async def create_class(
    cls: ClassCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_class = Class(**cls.model_dump())
    db.add(db_class)
    db.commit()
    db.refresh(db_class)

    return ClassResponse(
        id=db_class.id,
        class_name=db_class.class_name,
        grade=db_class.grade,
        description=db_class.description,
        teacher_id=db_class.teacher_id,
        is_active=db_class.is_active,
        created_at=db_class.created_at,
        teacher_name=None,
        student_count=0,
    )


@router.put("/api/classes/{class_id}", response_model=ClassResponse)
async def update_class(
    class_id: int,
    cls_update: ClassUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    update_data = cls_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_class, key, value)

    db.commit()
    db.refresh(db_class)

    teacher_name = None
    if db_class.teacher_id:
        teacher = db.query(User).filter(User.id == db_class.teacher_id).first()
        if teacher:
            teacher_name = teacher.real_name or teacher.username

    student_count = db.query(User).filter(User.class_id == db_class.id).count()

    return ClassResponse(
        id=db_class.id,
        class_name=db_class.class_name,
        grade=db_class.grade,
        description=db_class.description,
        teacher_id=db_class.teacher_id,
        is_active=db_class.is_active,
        created_at=db_class.created_at,
        teacher_name=teacher_name,
        student_count=student_count,
    )


@router.delete("/api/classes/{class_id}")
async def delete_class(
    class_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if not db_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    has_students = db.query(User).filter(User.class_id == class_id).first()
    if has_students:
        raise HTTPException(status_code=400, detail="班级下有学生，无法删除")

    db.delete(db_class)
    db.commit()

    return {"message": "班级已删除"}


@router.get("/api/classes/{class_id}/students", response_model=list[UserResponse])
async def get_class_students(
    class_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cls = db.query(Class).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    students = db.query(User).filter(User.class_id == class_id, User.role == "student").all()

    result = []
    for student in students:
        result.append(
            UserResponse(
                id=student.id,
                username=student.username,
                role=student.role,
                email=student.email,
                real_name=student.real_name,
                student_id=student.student_id,
                teacher_id=student.teacher_id,
                class_id=student.class_id,
                is_active=student.is_active,
                created_at=student.created_at,
                updated_at=student.updated_at,
                class_name=cls.class_name,
            )
        )

    return result


@router.get("/api/classes/{class_id}/devices")
async def get_class_devices(
    class_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cls = db.query(Class).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    binds = db.query(ClassDeviceBind).filter(ClassDeviceBind.class_id == class_id).all()

    result = []
    for bind in binds:
        device = db.query(Device).filter(Device.id == bind.device_id).first()
        if device:
            result.append(
                {
                    "id": bind.id,
                    "bind_id": bind.id,
                    "device_id": device.id,
                    "device_name": device.device_name,
                    "status": device.status,
                    "pump_state": device.pump_state,
                    "fan_state": device.fan_state,
                    "light_state": device.light_state,
                    "last_seen": device.last_seen,
                }
            )

    return result


@router.post("/api/classes/{class_id}/devices/bind")
async def bind_class_device(
    class_id: int,
    bind_data: ClassDeviceBindCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    cls = db.query(Class).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    device = db.query(Device).filter(Device.id == bind_data.device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    existing = (
        db.query(ClassDeviceBind)
        .filter(ClassDeviceBind.class_id == class_id, ClassDeviceBind.device_id == bind_data.device_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="该设备已绑定到此班级")

    db_bind = ClassDeviceBind(class_id=class_id, device_id=bind_data.device_id)
    db.add(db_bind)
    db.commit()
    db.refresh(db_bind)

    return {
        "id": db_bind.id,
        "class_id": class_id,
        "class_name": cls.class_name,
        "device_id": device.id,
        "device_name": device.device_name,
        "created_at": db_bind.created_at,
    }


@router.delete("/api/classes/{class_id}/devices/unbind/{bind_id}")
async def unbind_class_device(
    class_id: int,
    bind_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    db_bind = (
        db.query(ClassDeviceBind)
        .filter(ClassDeviceBind.id == bind_id, ClassDeviceBind.class_id == class_id)
        .first()
    )
    if not db_bind:
        raise HTTPException(status_code=404, detail="绑定关系不存在")

    db.delete(db_bind)
    db.commit()

    return {"message": "设备已解绑"}


@router.get("/api/students/{student_id}/device")
async def get_student_device(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == "student" and current_user.id != student_id:
        raise HTTPException(status_code=403, detail="无权查看")

    student = db.query(User).filter(User.id == student_id).first()
    if not student or not student.class_id:
        raise HTTPException(status_code=404, detail="学生未分配班级")

    binds = db.query(ClassDeviceBind).filter(ClassDeviceBind.class_id == student.class_id).all()

    result = []
    for bind in binds:
        device = db.query(Device).filter(Device.id == bind.device_id).first()
        if device:
            result.append(
                {
                    "id": device.id,
                    "device_name": device.device_name,
                    "status": device.status,
                    "pump_state": device.pump_state,
                    "fan_state": device.fan_state,
                    "light_state": device.light_state,
                }
            )

    return result
